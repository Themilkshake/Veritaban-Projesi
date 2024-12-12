# python C:\Users\lenovo\Desktop\MEKATEK_MUY\pyqt6_yer_istasyonu\yer_istasyonu_3.py & python C:\Users\lenovo\Desktop\MEKATEK_MUY\pyqt6_yer_istasyonu\video_html\app.py
import sys
from openpyxl import Workbook, load_workbook
from PyQt5.QtSql import QSqlDatabase, QSqlQuery
from PyQt5.QtWidgets import QApplication,QMainWindow,QMessageBox,QTableWidgetItem,QFileDialog, QWidget, QPushButton, QVBoxLayout, QLabel, QLineEdit, QTextEdit, QComboBox
from PyQt5 import QtWidgets
from giris_UI import Ui_MainWindowGiris
from personel_ana_ekrani_UI import Ui_MainWindowPerAnaEkran
from musteri_ana_ekrani_UI import Ui_MainWindowMusAnaEkran
from personel_profil_ekrani_UI import Ui_MainWindowPerProfil
from musteri_profil_ekrani_UI import Ui_MainWindowMusProfil
from baglan_UI import Ui_MainWindowBaglan
import pyodbc
import os

#------------------------------------------------

SERVER_NAME = 'ALI'
DATABASE_NAME = 'journey_management3'


#------------------------------------------------
# Veritabanı bağlantı bilgilerini buraya girin
conn_string = (
    r'DRIVER={SQL Server};'
    fr'SERVER={SERVER_NAME};'
    fr'DATABASE={DATABASE_NAME};'
    r'UID=;'
    r'PWD=;'
)

connString = f'DRIVER={{SQL Server}};'\
            f'SERVER={SERVER_NAME};'\
            f'DATABASE={DATABASE_NAME}'

def createConnection():
    global connString

    global db
    db = QSqlDatabase.addDatabase('QODBC')
    db.setDatabaseName(connString)

    if db.open():
        print('SQL SERVER basarili bir sekilde acildi')
        return True
    else:
        print('SQ SERVER baglanti hatasi')
        return False
#--------------------------------------------------

# class WindowBaglan(QMainWindow, Ui_MainWindowBaglan):
#     def __init__(self):
#         super().__init__()
#         self.setupUi(self)
#         self.baglanButton.clicked.connect(self.seydi)
    
#     def seydi(self):
#         global SERVER_NAME, DATABASE_NAME, connString, conn_string
#         DATABASE_NAME=self.SERVERADI.text().strip()
#         SERVER_NAME= self.DATABASEADI.text().strip()

#         conn_string = (
#             r'DRIVER={SQL Server};'
#             fr'SERVER={SERVER_NAME};'
#             fr'DATABASE={DATABASE_NAME};'
#             r'UID=;'
#             r'PWD=;'
#         )

#         connString = f'DRIVER={{SQL Server}};'\
#                     f'SERVER={SERVER_NAME};'\
#                     f'DATABASE={DATABASE_NAME}'
        
#         try:
#             conn = pyodbc.connect(conn_string) 

#             self.bilgilendirme_label.setText("Başarılı. Giriş yapılıyor...")
#             winGiris.show()
#             self.close()

#         except pyodbc.Error as ex:
#             print(ex)
#             self.bilgilendirme_label.setText("Giriş yaparken hata alıyoruz. Bilgileri kontrol edip tekrar deneyin.")
#         finally:
#             conn.close()

class WindowGiris(QMainWindow, Ui_MainWindowGiris):
    def __init__(self):
        super().__init__()
        self.setupUi(self)
        self.silme_log()
        self.musGirisButton.clicked.connect(self.load_data_mus)
        self.perGirisButton.clicked.connect(self.load_data_per)

    def load_data_per(self):
        # personel için veri çekme kodu
        if createConnection():  # Bağlantı kurulmuyorsa hata gösterin
            query = QSqlQuery(db)
            query.exec_("SELECT username FROM Employee")
            results_perUsername = []
            while query.next():
                results_perUsername.append(query.value("username"))


            query.exec_("SELECT employeeID, Password FROM Employee")
            results_perSif = []
            results_perID = []
            while query.next():
                results_perSif.append(query.value("Password"))
                results_perID.append(query.value("employeeID"))

        if db.isOpen():
            db.close()

        perUser=self.perUsername.text()
        perSif=self.perSif.text()
        per_user_sayac = 0
        giris_kontrol_1=0
        for i in results_perUsername:
            print(per_user_sayac)
            if i == perUser and results_perSif[per_user_sayac] == perSif:

                per_ID = results_perID[per_user_sayac]  # Employee ID değerini al

                try:
                    if createConnection():  # Bağlantı kurulumunu kontrol et
                        query = QSqlQuery(db)

                        # Çalışan bilgilerini sorgula
                        query.exec_(f"SELECT * FROM Employee WHERE employeeID = {per_ID}")

                        if query.next():
                            # Çalışan bilgilerini al
                            employeeID = query.value("employeeID")
                            firstName = query.value("firstName")
                            lastName = query.value("lastName")
                            phone = query.value("phone")
                            department = query.value("department")
                            username = query.value("username")
                            password = query.value("Password")
                            authorizationID = query.value("authorizationID")

                            # Veritabanına ekle
                            insert_query = QSqlQuery(db)
                            insert_query.prepare("""
                                INSERT INTO LogEmployee 
                                (employeeID, firstName, lastName, phone, department, username, password, authorizationID)
                                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                            """)
                            insert_query.addBindValue(employeeID)
                            insert_query.addBindValue(firstName)
                            insert_query.addBindValue(lastName)
                            insert_query.addBindValue(phone)
                            insert_query.addBindValue(department)
                            insert_query.addBindValue(username)
                            insert_query.addBindValue(password)
                            insert_query.addBindValue(authorizationID)

                            if insert_query.exec_():
                                print("Veriler LogEmployee tablosuna başarıyla eklendi.")
                            else:
                                print("Veri ekleme hatası:", insert_query.lastError().text())

                    if db.isOpen():
                        db.close()

                except Exception as ex:
                    print(f"Hata oluştu: {ex}")

                print("giriş başarili")
                self.close()
                winPerAnaEkran.show()
                giris_kontrol_1 = 0
                break
            else:
                giris_kontrol_1 = 1
            per_user_sayac +=1

        if giris_kontrol_1 == 1:
            QMessageBox.information(None, "Hata", "Kullanıcı adı yada parola hatalı.")

    def load_data_mus(self):
        # musteri için veri çekme kodu
        if createConnection():  # Bağlantı kurulmuyorsa hata gösterin
            query = QSqlQuery(db)
            query.exec_("SELECT email FROM Customer")
            results_musMail = []
            while query.next():
                results_musMail.append(query.value("email"))


            query.exec_("SELECT customerID, Password FROM Customer")
            results_musSif = []
            results_musID = []
            while query.next():
                results_musSif.append(query.value("Password"))
                results_musID.append(query.value("customerID"))
                
        if db.isOpen():
            db.close()

        #------------------
        musMail=self.musMail.text()
        musSif=self.musSif.text()
        mus_mail_sayac = 0
        giris_kontrol=0
        for i in results_musMail:
            print(results_musSif)
            if i == musMail and results_musSif[mus_mail_sayac] == musSif:

                mus_ID = results_musID[mus_mail_sayac]

                try:
                    if createConnection():  # Bağlantı kurulumunu kontrol et
                        query = QSqlQuery(db)

                        # Müşteri bilgilerini sorgula
                        query.exec_(f"SELECT * FROM Customer WHERE customerID = {mus_ID}")

                        if query.next():
                            # Müşteri bilgilerini al
                            customerID = query.value("customerID")
                            firstName = query.value("firstName")
                            lastName = query.value("lastName")
                            phone = query.value("phone")
                            email = query.value("email")
                            password = query.value("Password")
                            tcKimlikNo = query.value("tcKimlikNo")
                            authorizationID = query.value("authorizationID")

                            # Veritabanına ekle
                            insert_query = QSqlQuery(db)
                            insert_query.prepare("""
                                INSERT INTO LogCustomer 
                                (customerID, firstName, lastName, phone, email, password, tcKimlikNo, authorizationID)
                                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                            """)
                            insert_query.addBindValue(customerID)
                            insert_query.addBindValue(firstName)
                            insert_query.addBindValue(lastName)
                            insert_query.addBindValue(phone)
                            insert_query.addBindValue(email)
                            insert_query.addBindValue(password)
                            insert_query.addBindValue(tcKimlikNo)
                            insert_query.addBindValue(authorizationID)

                            if insert_query.exec_():
                                print("Veriler LogCustomer tablosuna başarıyla eklendi.")
                            else:
                                print("Veri ekleme hatası:", insert_query.lastError().text())

                    if db.isOpen():
                        db.close()

                except Exception as ex:
                    print(f"Hata oluştu: {ex}")


                print("giris başarili")

                self.close()
                winMusAnaEkran.show()

                giris_kontrol = 0
                break
            else:
                giris_kontrol = 1
            
            mus_mail_sayac +=1

        if giris_kontrol == 1:
            QMessageBox.information(None, "Hata", "Kullanıcı adı yada parola hatalı.")
        #********************

    def silme_log(self):
        try:
            if createConnection():  # Bağlantı kurulumunu kontrol et
                query = QSqlQuery(db)

                # LogCustomer tablosunu temizleme sorgusu
                delete_query = "TRUNCATE TABLE LogCustomer"
                if query.exec_(delete_query):
                    print("LogCustomer tablosu başarıyla temizlendi.")
                else:
                    print(f"Silme hatası: {query.lastError().text()}")


                delete_query2 = "TRUNCATE TABLE LogEmployee"
                if query.exec_(delete_query2):
                    print("LogEmployee tablosu başarıyla temizlendi.")
                else:
                    print(f"Silme hatası: {query.lastError().text()}")

            if db.isOpen():
                db.close()

        except Exception as ex:
            print(f"Bilinmeyen bir hata oluştu!\n{ex}")

class WindowPerAnaEkran(QMainWindow, Ui_MainWindowPerAnaEkran):
    def __init__(self):
        super().__init__()
        self.setupUi(self)

        self.yenile()
        self.cikisButton.clicked.connect(self.cikis)
        self.yenileButton.clicked.connect(self.yenile)
        self.profilButton.clicked.connect(self.profil)

        self.SeferTanimlaButton.clicked.connect(self.seferKontrol)
        self.aracTanimlaButton.clicked.connect(self.aracKontrol)
        self.musTanimlaButton.clicked.connect(self.musteri_ekleme)
        self.surucuTanimlaButton.clicked.connect(self.surucu_ekleme)
        self.perKaydetButton.clicked.connect(self.personel_ekleme)
        
        
        self.seferiIptalEtButton.clicked.connect(self.sefer_silme)
        self.musSilButton.clicked.connect(self.musteri_silme)
        self.surucuSilButton.clicked.connect(self.surucu_silme)
        self.perSilButton.clicked.connect(self.personel_silme)
        self.aracSilButton.clicked.connect(self.arac_silme)

        self.seferSorgulamafiltreButton.clicked.connect(self.sefer_sorgulama)
        self.aracSorgulamafiltreButton.clicked.connect(self.arac_sorgulama)
        self.musSorgulamafiltreButton.clicked.connect(self.musteri_sorgulama)
        self.surucuSorgulamafiltreButton.clicked.connect(self.surucu_sorgulama)

        self.excelSeferButton.clicked.connect(self.sefer_filtreleme)
        self.excelAracButton.clicked.connect(self.arac_filtreleme)
        self.excelMusteriButton.clicked.connect(self.musteri_filtreleme)
        self.excelSurucuButton.clicked.connect(self.surucu_filtreleme)
        self.excelSilinmisMusteriButton.clicked.connect(self.silinmis_musteri_filtreleme)
        self.exceldenVeritabaninaKaydetButton.clicked.connect(self.driver_excel_veri_aktar)

        self.yenileSilinmisButton.clicked.connect(self.silinmis_yenile)
        self.yedek_giris_button.clicked.connect(self.yedek_giris)
        self.yedegiGeriYukleButton.clicked.connect(self.yedegi_geri_yukle)
        self.SQLSorguCalistirButton.clicked.connect(self.SQLSorguyuCalistir)
        self.guncelleYenileButton.clicked.connect(self.musteri_veri_cekme_fonksiyonu)
        self.guncelleSecButton.clicked.connect(self.secili_satirdaki_verileri_linedite_cekme_fonksiyonu)
        self.guncelleGuncelleButton.clicked.connect(self.lineeditteki_verileri_guncelleme_fonksiyonu)
        self.BiletlerButton.clicked.connect(self.biletler_yukle)

        self.kilidiAcButton.clicked.connect(self.kullanici_yetki)
        self.yetkiButton.clicked.connect(self.yetki_ver)
        self.SQLtextEdit.setPlaceholderText("SQL sorgusunu buraya yazın...")

    def biletler_yukle(self):
        try:
            conn = pyodbc.connect(conn_string)
            cursor = conn.cursor()

            # Ticket tablosundan verileri çekme
            cursor.execute("SELECT ticketID, seatNumber, purchaseDate, amount, customerID, journeyID FROM [journey_management3].[dbo].[Ticket]")
            for row in cursor.fetchall():
                # tableWidget'e veriler yüklenecek
                # Burada her bir Ticket kaydını tableWidget'a ekleyebilirsiniz
                self.tableWidget.insertRow(self.tableWidget.rowCount())  # Yeni bir satır ekle
                self.tableWidget.setItem(self.tableWidget.rowCount() - 1, 0, QtWidgets.QTableWidgetItem(str(row.ticketID)))
                self.tableWidget.setItem(self.tableWidget.rowCount() - 1, 1, QtWidgets.QTableWidgetItem(str(row.seatNumber)))
                self.tableWidget.setItem(self.tableWidget.rowCount() - 1, 2, QtWidgets.QTableWidgetItem(str(row.purchaseDate)))
                self.tableWidget.setItem(self.tableWidget.rowCount() - 1, 3, QtWidgets.QTableWidgetItem(str(row.amount)))
                self.tableWidget.setItem(self.tableWidget.rowCount() - 1, 4, QtWidgets.QTableWidgetItem(str(row.customerID)))
                self.tableWidget.setItem(self.tableWidget.rowCount() - 1, 5, QtWidgets.QTableWidgetItem(str(row.journeyID)))

        except pyodbc.Error as ex:
            print(ex)

        finally:
            conn.close()


    def kullanici_yetki(self):
        try:
            conn = pyodbc.connect(conn_string)
            cursor = conn.cursor()

            cursor.execute("SELECT authorizationID FROM LogEmployee")
            for row in cursor.fetchall():
                self.tabWidget.setEnabled(True)
                
                if row.authorizationID != 1:
                    self.tabWidget.setTabEnabled(0, True)
                    self.tabWidget.setTabEnabled(1, True)
                    self.tabWidget.setTabEnabled(2, True)
                    self.tabWidget.setTabEnabled(3, False)
                    self.tabWidget.setTabEnabled(4, True)
                    self.tabWidget.setTabEnabled(5, False)
                    self.tabWidget.setTabEnabled(6, True)
                    self.tabWidget.setTabEnabled(7, False)
                    self.tabWidget.setTabEnabled(8, False)
                    self.tabWidget.setTabEnabled(10, False)
                    self.tabWidget.setTabEnabled(12, False)
                else:
                    self.tabWidget.setTabEnabled(0, True)
                    self.tabWidget.setTabEnabled(1, True)
                    self.tabWidget.setTabEnabled(2, True)
                    self.tabWidget.setTabEnabled(3, True)
                    self.tabWidget.setTabEnabled(4, True)
                    self.tabWidget.setTabEnabled(5, True)
                    self.tabWidget.setTabEnabled(6, True)
                    self.tabWidget.setTabEnabled(7, True)
                    self.tabWidget.setTabEnabled(8, True)
                    self.tabWidget.setTabEnabled(9, True)

                
        except pyodbc.Error as ex:
            print("SQL HATA VERDİ")

        finally:
            conn.close()


    def profil(self):
        winPerProfil.show()
        self.close()

    def silinmis_yenile(self):
        try:
            conn = pyodbc.connect(conn_string)
            cursor = conn.cursor()

            # SQL Sorgusu
            cursor.execute("""
                SELECT   
                    customerID, 
                    firstName, 
                    lastName, 
                    phone, 
                    email,  
                    tcKimlikNo, 
                    authorizationID 
                FROM dbo.DeletedCustomer
            """)

            rows = cursor.fetchall()

            # TableWidget Ayarları
            self.TableWidgetSilinmis.setRowCount(len(rows))

            # Verileri TableWidget'e Ekle
            for row_idx, row in enumerate(rows):
                for col_idx, col in enumerate(row):
                    self.TableWidgetSilinmis.setItem(
                        row_idx, col_idx, QTableWidgetItem(str(col))
                    )

        except pyodbc.Error as ex:
            print(f"Hata oluştu: {ex}")

        finally:
            conn.close()


    def yenile(self):
        self.comboboxSilinecekSefer.clear()
        self.comboboxSilinecekArac.clear()
        self.comboboxSilinecekMusteri.clear()
        self.comboboxSilinecekSurucu.clear()
        self.comboboxSilinecekPersonel.clear()

        self.comboboxSilinecekSefer.addItem(" ")
        self.comboboxSilinecekArac.addItem(" ")
        self.comboboxSilinecekMusteri.addItem(" ")
        self.comboboxSilinecekSurucu.addItem(" ")
        self.comboboxSilinecekPersonel.addItem(" ")

        self.comboboxKalkisSaati.clear()
        self.comboboxVarisSaati.clear()
        self.comboboxKalkisYeriFiltre.clear()
        self.comboboxVarisYeriFiltre.clear()
        self.comboboxAracMarka_2.clear()
        self.comboboxAracModel_2.clear()
        self.comboboxAracPlaka.clear()
        self.comboboxYetkiPersonel.clear()
    

        self.comboboxKalkisSaati.addItem(" ")
        self.comboboxVarisSaati.addItem(" ")
        self.comboboxKalkisYeriFiltre.addItem(" ") 
        self.comboboxVarisYeriFiltre.addItem(" ")
        self.comboboxAracMarka_2.addItem(" ") 
        self.comboboxAracModel_2.addItem(" ") 
        self.comboboxAracPlaka.addItem(" ") 
        self.comboboxYetkiPersonel.addItem(" ")


        self.aracSurucuGetirSql()
        self.aracOzellikGetirSql()
        self.SeferveAracSrogulamaFiltrelemeGetir()
        self.silinecek_arac_musteri_surucu_getir_sql()
        self.silinmis_yenile()
        self.yetki_yenile()

    def aracSurucuGetirSql(self):
        try:
            conn = pyodbc.connect(conn_string)
            cursor = conn.cursor()

            cursor.execute("SELECT driverID, firstName, lastName FROM Driver")
            for row in cursor.fetchall():
                self.comboboxSurucuSecimi.addItem(f"{row.driverID} {row.firstName} {row.lastName}")
            
            cursor.execute("SELECT busID, marka FROM Bus")
            for row in cursor.fetchall():
                self.comboboxAracSecimi.addItem(f"{row.busID} {row.marka}")  
        except pyodbc.Error as ex:
            print(ex)

        finally:
            conn.close()

    def aracOzellikGetirSql(self):
        try:
            conn = pyodbc.connect(conn_string)
            cursor = conn.cursor()

            cursor.execute("SELECT bustypeID, Feature FROM BusType")
            for row in cursor.fetchall():
                self.comboboxAracOzellikleri.addItem(f"{row.bustypeID} {row.Feature}")   
        except pyodbc.Error as ex:
            print(ex)

        finally:
            conn.close()

    def SeferveAracSrogulamaFiltrelemeGetir(self):
        try:
            conn = pyodbc.connect(conn_string)
            cursor = conn.cursor()

            cursor.execute("SELECT departureDate, departureCity, arrivalDate, arrivalCity FROM Journey")
            for row in cursor.fetchall():
                self.comboboxKalkisSaati.addItem(f"{row.departureDate}")
                self.comboboxVarisSaati.addItem(f"{row.arrivalDate}")
                self.comboboxKalkisYeriFiltre.addItem(f"{row.departureCity}")
                self.comboboxVarisYeriFiltre.addItem(f"{row.arrivalCity}") 
            
            cursor.execute("SELECT marka, model, numberPlate FROM Bus")
            for row in cursor.fetchall():
                self.comboboxAracMarka_2.addItem(f"{row.marka}") 
                self.comboboxAracModel_2.addItem(f"{row.model}") 
                self.comboboxAracPlaka.addItem(f"{row.numberPlate}") 

        except pyodbc.Error as ex:
            print(ex)

        finally:
            conn.close()


    def cikis(self):
        reply = QMessageBox.question(self, '!', 'Çıkmak istediğinize emin misiniz?', QMessageBox.Yes | QMessageBox.No)
        if reply == QMessageBox.Yes:
            app.quit()

    def seferKontrol(self):
        reply = QMessageBox.question(self, '!', 'Seferi kaydetmek istadiğinizden emin misiniz?', QMessageBox.Yes | QMessageBox.No)
        if reply == QMessageBox.Yes:
            self.sefer_tanimla()
        
    def sefer_tanimla(self):
        kalkisNoktasi=self.kalkis_noktasi.text()
        varisNoktasi=self.varis_noktasi.text()
        kalkisTarih=self.kalkis_tarih.text()
        varisTarih=self.varis_tarih.text()
        #sürücü ve araç seçimi kodu burada olacak.

        AracSecimi = self.comboboxAracSecimi.currentText()
        SurucuSecimi = self.comboboxSurucuSecimi.currentText()
        driverID = SurucuSecimi.split(" ", 1)[0]  # Boşluktan ayırma
        busID= AracSecimi.split(" ", 1)[0]

        if kalkisNoktasi and varisNoktasi:
            try:
                conn = pyodbc.connect(conn_string)
                cursor = conn.cursor()
                query = """
                INSERT INTO Journey (departureCity, arrivalCity, departureDate, arrivalDate, busID, driverID)
                VALUES (?, ?, ?, ?, ?, ?)
                """
                # Kullanıcı verilerini sorguya bağla ve çalıştır
                cursor.execute(query, (kalkisNoktasi, varisNoktasi, kalkisTarih, varisTarih, busID, driverID))
                conn.commit()

            except pyodbc.Error as ex:
                print(ex)

            finally:
                conn.close()
        else:
            QMessageBox.warning(self, "Uyarı", "Kalkış Noktası ve Varış Noktası giriniz.")

        #Buton ayarlanacak.

    def aracKontrol(self):
        reply = QMessageBox.question(self, '!', 'Aracı kaydetmek istadiğinizden emin misiniz?', QMessageBox.Yes | QMessageBox.No)
        if reply == QMessageBox.Yes:
            self.arac_tanimla()
    

    def arac_tanimla(self):
        aracKapasite=self.aracKapasite.text()
        aracPlakaKodu=self.aracPlakaKodu.text()
        aracMarka=self.aracMarka.text()
        aracModel=self.aracModel.text()
        AracOzellikleri = self.comboboxAracOzellikleri.currentText()
        bustypeID, Feature = AracOzellikleri.split(" ",1)  # Boşluktan ayırma

        if aracKapasite and aracPlakaKodu and aracMarka and aracModel and bustypeID:
            try:
                # Veritabanı bağlantısı
                conn = pyodbc.connect(conn_string)
                cursor = conn.cursor()

                # SQL sorgusu
                query = """
                INSERT INTO Bus (capacity, numberPlate, marka, bustypeID, model)
                VALUES (?, ?, ?, ?, ?)
                """

                # Kullanıcı verilerini sorguya bağla ve çalıştır
                cursor.execute(query, (aracKapasite, aracPlakaKodu, aracMarka, bustypeID, aracModel))
                conn.commit()
                
                # Başarılı mesajı
                QMessageBox.information(self, "Başarılı", "Araç başarıyla eklendi.")
            
            except pyodbc.Error as ex:
                # Hata mesajı
                print("Veritabanı Hatası:", ex)
                QMessageBox.critical(self, "Hata", "Araç eklenirken bir hata oluştu.")
            
            finally:
                # Bağlantıyı kapatma
                conn.close()
        else:
            # Eksik bilgi uyarısı
            QMessageBox.warning(self, "Uyarı", "Lütfen tüm bilgileri doldurun.")
    
    def personel_ekleme(self):
        perAd=self.perAd.text()
        perSoyad=self.perSoyad.text()
        perTelNo=self.perTelNo.text()
        perDepartman=self.perDepartman.text()
        perUsername=self.perUsername.text()
        perSif=self.perSif.text()
        musSifTekrar=self.perSifTekrar.text()
        authorization=2

        #********************
        sayac=0
        try:
            conn = pyodbc.connect(conn_string)
            cursor = conn.cursor()

            cursor.execute("SELECT employeeID, username FROM Employee")
            
            for row in cursor.fetchall():
                if row.username == perUsername:
                    sayac=1
                else:
                    sayac=0
              
        except pyodbc.Error as ex:
            print(ex)

        finally:
            conn.close()

        #********************


        if (sayac != 1 and perAd and perSoyad and perTelNo and perDepartman 
            and perUsername and perSif and musSifTekrar and perSif == musSifTekrar):
            try:
                # Veritabanı bağlantısı
                conn = pyodbc.connect(conn_string)
                cursor = conn.cursor()

                # SQL sorgusu
                query = """
                    INSERT INTO Employee (firstName, lastName, phone, department, username, Password, authorizationID)
                    VALUES (?, ?, ?, ?, ?, ?, ?)
                    """

                # Kullanıcı verilerini sorguya bağla ve çalıştır
                cursor.execute(query, (perAd, perSoyad, perTelNo, perDepartman, perUsername, perSif, authorization))
                conn.commit()

                # Başarılı mesajı
                QMessageBox.information(self, "Başarılı", "Personel başarıyla eklendi.")
            
            except pyodbc.Error as ex:
                # Hata mesajı
                print("Veritabanı Hatası:", ex)
                QMessageBox.critical(self, "Hata", "Personel eklenirken bir hata oluştu.")
            
            finally:
                # Bağlantıyı kapatma
                conn.close()
        else:
            # Eksik bilgi veya şifre uyumsuzluğu uyarısı
            if perSif != musSifTekrar:
                QMessageBox.warning(self, "Uyarı", "Şifreler uyuşmuyor.")
            else:
                QMessageBox.warning(self, "Uyarı", "Lütfen tüm bilgileri doldurun. Kullanıcı ismi başka bir kullanıcı tarafından kullanılıyor olabilir.")

    def yetki_yenile(self):
        try:
            conn = pyodbc.connect(conn_string)
            cursor = conn.cursor()
            cursor.execute("SELECT employeeID, firstName, lastName, phone, department, username, authorizationID FROM Employee")
            for row in cursor.fetchall():
                self.comboboxYetkiPersonel.addItem(
                    f"{row.employeeID} | İsim Soyisim: {row.firstName} {row.lastName} | Telefon Numarası: {row.phone} | Departman: {row.department} | Kullanıcı Adı: {row.username} | Yetki ID: {row.authorizationID}")

        except pyodbc.Error as ex:
            print(ex)

        finally:
            conn.close()

    def yetki_ver(self):
        SilinecekPersonel = self.comboboxYetkiPersonel.currentText()
        employeeID = SilinecekPersonel.split(" ", 1)[0]
        authorizationID = self.comboboxYetkiDerece.currentText()

        if self.comboboxYetkiPersonel.currentText() != " " and self.comboboxYetkiDerece.currentText() != "":
            response = QMessageBox.question(self, "Koşul Sağlandı", "Personeli yetkisini değişirmek istediğinizden emin misiniz?",
                                            QMessageBox.Yes | QMessageBox.No)
            if response == QMessageBox.Yes:
                
                update_query = """
                UPDATE Employee
                SET authorizationID = ?
                WHERE employeeID = ?
                """

                try:
                    conn = pyodbc.connect(conn_string)
                    cursor = conn.cursor()
                    cursor.execute(update_query,(authorizationID,employeeID))
                    conn.commit()
                    QMessageBox.information(self, "Başarılı", "Personel Yetkisi değiştirilmiştir")
                    
                except pyodbc.Error as ex:
                    QMessageBox.warning(self, "Hata", "Veritabanında hata oldu tekrar deneyiniz")

                finally:
                    conn.close()
                
        else:
            QMessageBox.warning(self, "Hata", "Seçim yapınız")



    def musteri_ekleme(self):

        musAd=self.musAd.text()
        musSoyad=self.musSoyad.text()
        musKimlik=self.musKimlik.text()
        musTelNo=self.musTelNo.text()
        musMail=self.musMail.text()
        musSif=self.musSif.text()
        musSifTekrar=self.musSifTekrar.text()
        authorization=3

        #********************
        sayac=0
        try:
            conn = pyodbc.connect(conn_string)
            cursor = conn.cursor()

            cursor.execute("SELECT customerID, email FROM Customer")
            
            for row in cursor.fetchall():
                if row.email == musMail:
                    sayac=1
                else:
                    sayac=0
              
        except pyodbc.Error as ex:
            print(ex)

        finally:
            conn.close()

        #********************




        if (sayac != 1 and musAd and musSoyad and musKimlik and musTelNo and musMail 
            and musSif and musSifTekrar and musSif == musSifTekrar):
            try:
                # Veritabanı bağlantısı
                conn = pyodbc.connect(conn_string)
                cursor = conn.cursor()

                # SQL sorgusu
                query = """
                INSERT INTO Customer (firstName, lastName, phone, email, Password, tcKimlikNo, authorizationID)
                VALUES (?, ?, ?, ?, ?, ?, ?)
                """

                # Kullanıcı verilerini sorguya bağla ve çalıştır
                cursor.execute(query, (musAd, musSoyad, musTelNo, musMail, musSif, musKimlik, authorization))
                conn.commit()

                # Başarılı mesajı
                QMessageBox.information(self, "Başarılı", "Müşteri başarıyla eklendi.")
            
            except pyodbc.Error as ex:
                # Hata mesajı
                print("Veritabanı Hatası:", ex)
                QMessageBox.critical(self, "Hata", "Müşteri eklenirken bir hata oluştu.")
            
            finally:
                # Bağlantıyı kapatma
                conn.close()
        else:
            # Eksik bilgi veya şifre uyumsuzluğu uyarısı
            if musSif != musSifTekrar:
                QMessageBox.warning(self, "Uyarı", "Şifreler uyuşmuyor.")
            else:
                QMessageBox.warning(self, "Uyarı", "Lütfen tüm bilgileri doldurun. Email başka bir kullanıcı tarafından kullanılıyor olabilir.")



    def surucu_ekleme(self):
        # Kullanıcıdan alınacak veriler
        driverFirstName = self.surucuAd.text()  # QLineEdit'den ad bilgisi
        driverLastName = self.surucuSoyad.text()  # QLineEdit'den soyad bilgisi
        driverPhone = self.surucuTelNo.text()  # QLineEdit'den telefon numarası
        driverLicenseNumber = self.surucuLisansNo.text()  # QLineEdit'den ehliyet numarası

        # Boş alan kontrolü
        if driverFirstName and driverLastName and driverPhone and driverLicenseNumber:
            try:
                # Veritabanı bağlantısı
                conn = pyodbc.connect(conn_string)
                cursor = conn.cursor()

                # SQL sorgusu
                query = """
                INSERT INTO Driver (firstName, lastName, phone, licenseNumber)
                VALUES (?, ?, ?, ?)
                """

                # Kullanıcı verilerini sorguya bağla ve çalıştır
                cursor.execute(query, (driverFirstName, driverLastName, driverPhone, driverLicenseNumber))
                conn.commit()

                # Başarılı mesajı
                QMessageBox.information(self, "Başarılı", "Sürücü başarıyla eklendi.")
            
            except pyodbc.Error as ex:
                # Hata mesajı
                print("Veritabanı Hatası:", ex)
                QMessageBox.critical(self, "Hata", "Sürücü eklenirken bir hata oluştu.")
            
            finally:
                conn.close()
        else:
            QMessageBox.warning(self, "Uyarı", "Lütfen tüm bilgileri doldurun.")

    def sefer_sorgulama(self):
        departureDatee = self.comboboxKalkisSaati.currentText()
        arrivalDatee = self.comboboxVarisSaati.currentText()
        departureCityy = self.comboboxKalkisYeriFiltre.currentText()
        arrivalCityy = self.comboboxVarisYeriFiltre.currentText()
        
        query = "SELECT * FROM Journey WHERE 1=1"
        params = []

        # Filtreleri ekle
        if departureCityy != " ":
            query += " AND departureCity = ?"
            params.append(departureCityy)

        if arrivalCityy != " ":
            query += " AND arrivalCity = ?"
            params.append(arrivalCityy)

        if departureDatee != " ":
            query += " AND departureDate = ?"
            params.append(departureDatee)

        if arrivalDatee != " ":
            query += " AND arrivalDate = ?"
            params.append(arrivalDatee)
        
        print(params)
        try:
            # Veritabanına bağlan ve sorguyu çalıştır
            conn = pyodbc.connect(conn_string)
            cursor = conn.cursor()
            cursor.execute(query,params)
            rows = cursor.fetchall()
            # TableWidget'i temizle
            self.tableWidgetSeferler.setRowCount(0)

            # Sorgu sonuçlarını TableWidget'e ekle
            for row_number, row_data in enumerate(rows):
                
                self.tableWidgetSeferler.insertRow(row_number)
                for column_number, data in enumerate(row_data):

                    item = QTableWidgetItem(str(data) if data is not None else "") 
                    self.tableWidgetSeferler.setItem(row_number, column_number, item)
        except pyodbc.Error as ex:
            print("Veritabanı Hatası:", ex)
            QMessageBox.critical(self, "Hata", "Veritabanı bağlantı hatası!")

        finally:
            conn.close()

    def sefer_filtreleme(self):

        dosya_yolu, _ = QFileDialog.getSaveFileName(
            self, "Excel Dosyasını Kaydet", "", "Excel Dosyası (*.xlsx)"
        )
        
        if not dosya_yolu:
            return  # Dosya seçilmedi

        try:
            # Excel çalışma kitabı oluştur
            wb = Workbook()
            ws = wb.active
            ws.title = "Seferler"

            # TableWidget başlıklarını yaz
            kolon_sayisi = self.tableWidgetSeferler.columnCount()
            satir_sayisi = self.tableWidgetSeferler.rowCount()

            # Başlıkları Excel'e ekle
            for kolon in range(kolon_sayisi):
                baslik = self.tableWidgetSeferler.horizontalHeaderItem(kolon).text()
                ws.cell(row=1, column=kolon + 1, value=baslik)

            # TableWidget verilerini Excel'e ekle
            for satir in range(satir_sayisi):
                for kolon in range(kolon_sayisi):
                    hucre_verisi = self.tableWidgetSeferler.item(satir, kolon)
                    ws.cell(row=satir + 2, column=kolon + 1, value=hucre_verisi.text() if hucre_verisi else "")

            # Dosyayı kaydet
            wb.save(dosya_yolu)
            QMessageBox.information(self, "Başarılı", "Veriler Excel dosyasına başarıyla aktarıldı!")
        
        except Exception as ex:
            QMessageBox.critical(self, "Hata", f"Dosya kaydedilirken hata oluştu!\n{ex}")


    def arac_sorgulama(self):
        plaka_2 = self.comboboxAracPlaka.currentText()
        marka_2 = self.comboboxAracMarka_2.currentText()
        model_2 = self.comboboxAracModel_2.currentText()
        
        query = "SELECT busID, capacity, numberPlate, marka, model FROM Bus WHERE 1=1"
        params = []

        # Filtreleri ekle
        if plaka_2 != " ":
            query += " AND numberPlate = ?"
            params.append(plaka_2)

        if marka_2 != " ":
            query += " AND marka = ?"
            params.append(marka_2)

        if model_2 != " ":
            query += " AND model = ?"
            params.append(model_2)

        
        try:
            # Veritabanına bağlan ve sorguyu çalıştır
            conn = pyodbc.connect(conn_string)
            cursor = conn.cursor()
            cursor.execute(query,params)
            rows = cursor.fetchall()
            # TableWidget'i temizle
            self.aracSorgulama_tableWidget.setRowCount(0)

            # Sorgu sonuçlarını TableWidget'e ekle
            for row_number, row_data in enumerate(rows):
                
                self.aracSorgulama_tableWidget.insertRow(row_number)
                for column_number, data in enumerate(row_data):

                    item = QTableWidgetItem(str(data) if data is not None else "") 
                    self.aracSorgulama_tableWidget.setItem(row_number, column_number, item)
        except pyodbc.Error as ex:
            print("Veritabanı Hatası:", ex)
            QMessageBox.critical(self, "Hata", "Veritabanı bağlantı hatası!")

        finally:
            conn.close()

    def arac_filtreleme(self):

        dosya_yolu, _ = QFileDialog.getSaveFileName(
            self, "Excel Dosyasını Kaydet", "", "Excel Dosyası (*.xlsx)"
        )
        
        if not dosya_yolu:
            return  # Dosya seçilmedi

        try:
            # Excel çalışma kitabı oluştur
            wb = Workbook()
            ws = wb.active
            ws.title = "Seferler"

            # TableWidget başlıklarını yaz
            kolon_sayisi = self.aracSorgulama_tableWidget.columnCount()
            satir_sayisi = self.aracSorgulama_tableWidget.rowCount()

            # Başlıkları Excel'e ekle
            for kolon in range(kolon_sayisi):
                baslik = self.aracSorgulama_tableWidget.horizontalHeaderItem(kolon).text()
                ws.cell(row=1, column=kolon + 1, value=baslik)

            # TableWidget verilerini Excel'e ekle
            for satir in range(satir_sayisi):
                for kolon in range(kolon_sayisi):
                    hucre_verisi = self.aracSorgulama_tableWidget.item(satir, kolon)
                    ws.cell(row=satir + 2, column=kolon + 1, value=hucre_verisi.text() if hucre_verisi else "")

            # Dosyayı kaydet
            wb.save(dosya_yolu)
            QMessageBox.information(self, "Başarılı", "Veriler Excel dosyasına başarıyla aktarıldı!")
        
        except Exception as ex:
            QMessageBox.critical(self, "Hata", f"Dosya kaydedilirken hata oluştu!\n{ex}")
    
    def musteri_sorgulama(self):

        # ComboBox değerlerini al
        firstName = self.musSorgulama_MusAd.text() or None
        lastName = self.musSorgulama_MusSoyad.text() or None
        phone = self.musSorgulama_MusTelNo.text() or None
        email = self.musSorgulama_MusMail.text() or None

        try:
            # Veritabanına bağlan
            conn = pyodbc.connect(conn_string)
            cursor = conn.cursor()

            # Saklı yordamı çağır
            cursor.execute("""
                EXEC SearchCustomers 
                    @firstName = ?, 
                    @lastName = ?, 
                    @phone = ?, 
                    @email = ?
            """, firstName, lastName, phone, email)

            rows = cursor.fetchall()
            
            # TableWidget'i temizle
            self.musSorgulama_tableWidget.setRowCount(0)

            # Sonuçları TableWidget'e ekle
            for row_number, row_data in enumerate(rows):
                print(row_data)
                self.musSorgulama_tableWidget.insertRow(row_number)
                for column_number, data in enumerate(row_data):
                    item = QTableWidgetItem(str(data) if data is not None else "")
                    self.musSorgulama_tableWidget.setItem(row_number, column_number, item)

        except pyodbc.Error as ex:
            print("Veritabanı Hatası:", ex)
            QMessageBox.critical(self, "Hata", "Veritabanı bağlantı hatası!")

        finally:
            conn.close()

    def musteri_filtreleme(self):

        dosya_yolu, _ = QFileDialog.getSaveFileName(
            self, "Excel Dosyasını Kaydet", "", "Excel Dosyası (*.xlsx)"
        )
        
        if not dosya_yolu:
            return  # Dosya seçilmedi

        try:
            # Excel çalışma kitabı oluştur
            wb = Workbook()
            ws = wb.active
            ws.title = "Seferler"

            # TableWidget başlıklarını yaz
            kolon_sayisi = self.musSorgulama_tableWidget.columnCount()
            satir_sayisi = self.musSorgulama_tableWidget.rowCount()

            # Başlıkları Excel'e ekle
            for kolon in range(kolon_sayisi):
                baslik = self.musSorgulama_tableWidget.horizontalHeaderItem(kolon).text()
                ws.cell(row=1, column=kolon + 1, value=baslik)

            # TableWidget verilerini Excel'e ekle
            for satir in range(satir_sayisi):
                for kolon in range(kolon_sayisi):
                    hucre_verisi = self.musSorgulama_tableWidget.item(satir, kolon)
                    ws.cell(row=satir + 2, column=kolon + 1, value=hucre_verisi.text() if hucre_verisi else "")

            # Dosyayı kaydet
            wb.save(dosya_yolu)
            QMessageBox.information(self, "Başarılı", "Veriler Excel dosyasına başarıyla aktarıldı!")
        
        except Exception as ex:
            QMessageBox.critical(self, "Hata", f"Dosya kaydedilirken hata oluştu!\n{ex}")


    def silinmis_musteri_filtreleme(self):

        dosya_yolu, _ = QFileDialog.getSaveFileName(
            self, "Excel Dosyasını Kaydet", "", "Excel Dosyası (*.xlsx)"
        )
        
        if not dosya_yolu:
            return  # Dosya seçilmedi

        try:
            # Excel çalışma kitabı oluştur
            wb = Workbook()
            ws = wb.active
            ws.title = "Seferler"

            # TableWidget başlıklarını yaz
            kolon_sayisi = self.TableWidgetSilinmis.columnCount()
            satir_sayisi = self.TableWidgetSilinmis.rowCount()

            # Başlıkları Excel'e ekle
            for kolon in range(kolon_sayisi):
                baslik = self.TableWidgetSilinmis.horizontalHeaderItem(kolon).text()
                ws.cell(row=1, column=kolon + 1, value=baslik)

            # TableWidget verilerini Excel'e ekle
            for satir in range(satir_sayisi):
                for kolon in range(kolon_sayisi):
                    hucre_verisi = self.TableWidgetSilinmis.item(satir, kolon)
                    ws.cell(row=satir + 2, column=kolon + 1, value=hucre_verisi.text() if hucre_verisi else "")

            # Dosyayı kaydet
            wb.save(dosya_yolu)
            QMessageBox.information(self, "Başarılı", "Veriler Excel dosyasına başarıyla aktarıldı!")
        
        except Exception as ex:
            QMessageBox.critical(self, "Hata", f"Dosya kaydedilirken hata oluştu!\n{ex}")


    def yedek_giris(self):
        winBackup.show()
        self.close()
        
    def yedegi_geri_yukle(self):
        winBackupRestore.show()
        self.close()



    def surucu_sorgulama(self):

        # ComboBox değerlerini al
        surucuAd = self.surucuSorgulama_SurucuAd.text() or None
        surucuSoyad = self.surucuSorgulama_SurucuSoyad.text() or None
        surucutelNo = self.surucuSorgulama_SurucuTelno.text() or None
        suruculinsansNo = self.surucuSorgulama_SurucuLisansNo.text() or None

        try:
            # Veritabanına bağlan
            conn = pyodbc.connect(conn_string)
            cursor = conn.cursor()

            # Saklı yordamı çağır
            cursor.execute("""
                EXEC SearchDrivers 
                    @firstName = ?, 
                    @lastName = ?, 
                    @phone = ?, 
                    @licenseNumber = ?
            """, surucuAd, surucuSoyad, surucutelNo, suruculinsansNo)

            # Sonuçları al
            rows = cursor.fetchall()

            # TableWidget'i temizle
            self.surucuSorgulama_tableWidget.setRowCount(0)

            # Sonuçları TableWidget'e ekle
            for row_number, row_data in enumerate(rows):
                self.surucuSorgulama_tableWidget.insertRow(row_number)
                for column_number, data in enumerate(row_data):
                    item = QTableWidgetItem(str(data) if data is not None else "")
                    self.surucuSorgulama_tableWidget.setItem(row_number, column_number, item)

        except pyodbc.Error as ex:
            print("Veritabanı Hatası:", ex)
            QMessageBox.critical(self, "Hata", "Veritabanı bağlantı hatası!")

        finally:
            conn.close()

    def surucu_filtreleme(self):
        dosya_yolu, _ = QFileDialog.getSaveFileName(
            self, "Excel Dosyasını Kaydet", "", "Excel Dosyası (*.xlsx)"
        )
        
        if not dosya_yolu:
            return  # Dosya seçilmedi

        try:
            # Excel çalışma kitabı oluştur
            wb = Workbook()
            ws = wb.active
            ws.title = "Seferler"

            # TableWidget başlıklarını yaz
            kolon_sayisi = self.surucuSorgulama_tableWidget.columnCount()
            satir_sayisi = self.surucuSorgulama_tableWidget.rowCount()

            # Başlıkları Excel'e ekle
            for kolon in range(kolon_sayisi):
                baslik = self.surucuSorgulama_tableWidget.horizontalHeaderItem(kolon).text()
                ws.cell(row=1, column=kolon + 1, value=baslik)

            # TableWidget verilerini Excel'e ekle
            for satir in range(satir_sayisi):
                for kolon in range(kolon_sayisi):
                    hucre_verisi = self.surucuSorgulama_tableWidget.item(satir, kolon)
                    ws.cell(row=satir + 2, column=kolon + 1, value=hucre_verisi.text() if hucre_verisi else "")

            # Dosyayı kaydet
            wb.save(dosya_yolu)
            QMessageBox.information(self, "Başarılı", "Veriler Excel dosyasına başarıyla aktarıldı!")
        
        except Exception as ex:
            QMessageBox.critical(self, "Hata", f"Dosya kaydedilirken hata oluştu!\n{ex}")

    def silinecek_arac_musteri_surucu_getir_sql(self):
        try:
            conn = pyodbc.connect(conn_string)
            cursor = conn.cursor()
            '''
            self.comboboxSilinecekSefer.addItem(" ")
            self.comboboxSilinecekArac.addItem(" ")
            self.comboboxSilinecekMusteri.addItem(" ")
            self.comboboxSilinecekSurucu.addItem(" ")
            '''
            
            cursor.execute("SELECT journeyID, departureCity, arrivalCity, departureDate, arrivalDate, busID, driverID FROM Journey")
            for row in cursor.fetchall():
                self.comboboxSilinecekSefer.addItem(f"{row.journeyID} Kalkış: {row.departureCity} | Varış: {row.arrivalCity} | Kalkış Saati: {row.departureDate} | Varış Saati: {row.arrivalDate}") 
            
            cursor.execute("SELECT busID, capacity, numberPlate, marka, bustypeID, model FROM Bus")
            for row in cursor.fetchall():
                self.comboboxSilinecekArac.addItem(f"{row.busID} | Kapasite: {row.capacity} | Plaka: {row.numberPlate} | Marka: {row.marka} | Modeli: {row.model}") 

            cursor.execute("SELECT driverID, firstName, lastName, phone, licenseNumber FROM Driver")
            for row in cursor.fetchall():
                self.comboboxSilinecekSurucu.addItem(f"{row.driverID} | İsim Soyisim: {row.firstName} {row.lastName} | Telefon Numarası: {row.phone} | Ehliyet Numarası: {row.licenseNumber}") 
            

            cursor.execute("SELECT customerID, firstName, lastName, phone, email, tcKimlikNo, authorizationID FROM Customer")
            for row in cursor.fetchall():
                self.comboboxSilinecekMusteri.addItem(f"{row.customerID} | İsim Soyisim: {row.firstName} {row.lastName} | Telefon Numarası: {row.phone} | E-mail: {row.email} | TC Kimlik No: {row.tcKimlikNo}")


            cursor.execute("SELECT employeeID, firstName, lastName, phone, department, username FROM Employee")
            for row in cursor.fetchall():
                self.comboboxSilinecekPersonel.addItem(
                    f"{row.employeeID} | İsim Soyisim: {row.firstName} {row.lastName} | Telefon Numarası: {row.phone} | Departman: {row.department} | Kullanıcı Adı: {row.username}")


        except pyodbc.Error as ex:
            print(ex)

        finally:
            conn.close()
        
    def sefer_silme(self):
        SilinecekSefer = self.comboboxSilinecekSefer.currentText()
        journeyID = SilinecekSefer.split(" ", 1)[0]

        if self.onayCheckBoxSefer.isChecked() and self.comboboxSilinecekSefer.currentText() != " ":
            response = QMessageBox.question(self, "Koşul Sağlandı", "Seferi silmek istediğinizden emin misiniz?",
                                            QMessageBox.Yes | QMessageBox.No)
            if response == QMessageBox.Yes:
                try:
                    conn = pyodbc.connect(conn_string)
                    cursor = conn.cursor()
                    cursor.execute("DELETE FROM Journey WHERE journeyID = ?",(journeyID,))
                    conn.commit()
                    QMessageBox.information(self, "Başarılı", "Sefer Silinmiştir")
                    
                except pyodbc.Error as ex:
                    QMessageBox.warning(self, "Hata", "Veritabanında hata oldu tekrar deneyiniz")

                finally:
                    conn.close()
                
        else:
            QMessageBox.warning(self, "Hata", "Seçim yapınız ve Kutucuğu işaretleyiniz")

    def personel_silme(self):

        SilinecekPersonel = self.comboboxSilinecekPersonel.currentText()
        employeeID = SilinecekPersonel.split(" ", 1)[0]

        if self.onayCheckBoxPersonel.isChecked() and self.comboboxSilinecekPersonel.currentText() != " ":
            response = QMessageBox.question(self, "Koşul Sağlandı", "Müşteriyi silmek istediğinizden emin misiniz?",
                                            QMessageBox.Yes | QMessageBox.No)
            if response == QMessageBox.Yes:
                try:
                    conn = pyodbc.connect(conn_string)
                    cursor = conn.cursor()
                    cursor.execute("DELETE FROM Employee WHERE employeeID = ?",(employeeID,))
                    conn.commit()
                    QMessageBox.information(self, "Başarılı", "Müşteri Silinmiştir")
                    
                except pyodbc.Error as ex:
                    QMessageBox.warning(self, "Hata", "Veritabanında hata oldu tekrar deneyiniz")

                finally:
                    conn.close()
                
        else:
            QMessageBox.warning(self, "Hata", "Seçim yapınız ve Kutucuğu işaretleyiniz")
        pass

    def musteri_silme(self):
        SilinecekMusteri = self.comboboxSilinecekMusteri.currentText()
        customerID = SilinecekMusteri.split(" ", 1)[0]

        if self.onayCheckBoxMusteri.isChecked() and self.comboboxSilinecekMusteri.currentText() != " ":
            response = QMessageBox.question(self, "Koşul Sağlandı", "Müşteriyi silmek istediğinizden emin misiniz?",
                                            QMessageBox.Yes | QMessageBox.No)
            if response == QMessageBox.Yes:
                try:
                    conn = pyodbc.connect(conn_string)
                    cursor = conn.cursor()
                    cursor.execute("DELETE FROM Customer WHERE customerID = ?",(customerID,))
                    conn.commit()
                    QMessageBox.information(self, "Başarılı", "Müşteri Silinmiştir")
                    
                except pyodbc.Error as ex:
                    QMessageBox.warning(self, "Hata", "Veritabanında hata oldu tekrar deneyiniz")

                finally:
                    conn.close()
                
        else:
            QMessageBox.warning(self, "Hata", "Seçim yapınız ve Kutucuğu işaretleyiniz")

    def surucu_silme(self):
        SilinecekSurucu = self.comboboxSilinecekSurucu.currentText()
        driverID = SilinecekSurucu.split(" ", 1)[0]

        if self.onayCheckBoxSurucu.isChecked() and self.comboboxSilinecekSurucu.currentText() != " ":
            response = QMessageBox.question(self, "Koşul Sağlandı", "Sürücü silmek istediğinizden emin misiniz?",
                                            QMessageBox.Yes | QMessageBox.No)
            if response == QMessageBox.Yes:
                try:
                    conn = pyodbc.connect(conn_string)
                    cursor = conn.cursor()
                    cursor.execute("DELETE FROM Driver WHERE driverID = ?",(driverID,))
                    conn.commit()
                    QMessageBox.information(self, "Başarılı", "Sürücü Silinmiştir")
                    
                except pyodbc.Error as ex:
                    QMessageBox.warning(self, "Hata", "Veritabanında hata oldu tekrar deneyiniz")

                finally:
                    conn.close()
                
        else:
            QMessageBox.warning(self, "Hata", "Seçim yapınız ve Kutucuğu işaretleyiniz")

    def arac_silme(self):
        SilinecekArac = self.comboboxSilinecekArac.currentText()
        busID = SilinecekArac.split(" ", 1)[0]

        if self.onayCheckBoxArac.isChecked() and self.comboboxSilinecekArac.currentText() != " ":
            response = QMessageBox.question(self, "Koşul Sağlandı", "Aracı silmek istediğinizden emin misiniz?",
                                            QMessageBox.Yes | QMessageBox.No)
            if response == QMessageBox.Yes:
                try:
                    conn = pyodbc.connect(conn_string)
                    cursor = conn.cursor()
                    cursor.execute("DELETE FROM Bus WHERE busID= ?",(busID,))
                    conn.commit()
                    QMessageBox.information(self, "Başarılı", "Araç Silinmiştir")
                    
                except pyodbc.Error as ex:
                    QMessageBox.warning(self, "Hata", "Veritabanında hata oldu tekrar deneyiniz")

                finally:
                    conn.close()
                
        else:
            QMessageBox.warning(self, "Hata", "Seçim yapınız ve Kutucuğu işaretleyiniz")


    def driver_excel_veri_aktar(self):
        # Excel dosyasını seç
        dosya_yolu, _ = QFileDialog.getOpenFileName(
            self, "Excel Dosyasını Seç", "", "Excel Dosyası (*.xlsx)"
        )

        if not dosya_yolu:
            return  # Dosya seçilmediyse çıkış yap
        
        try:
            # Excel dosyasını yükle
            wb = load_workbook(dosya_yolu)
            ws = wb.active

            # MSSQL veritabanına bağlan
            conn = pyodbc.connect(conn_string)
            cursor = conn.cursor()

            # Satırları oku ve veritabanına ekle
            for row in ws.iter_rows(min_row=2, values_only=True):  # Başlıkları atla
                firstName, lastName, phone, licenseNumber = row

                # SQL Sorgusu
                query = """
                INSERT INTO Driver (firstName, lastName, phone, licenseNumber)
                VALUES (?, ?, ?, ?)
                """
                params = (firstName, lastName, phone, licenseNumber)
                cursor.execute(query, params)

            conn.commit()  # Veritabanına işlemleri kaydet
            QMessageBox.information(self, "Başarılı", "Veriler başarıyla MSSQL'e aktarıldı!")

        except FileNotFoundError:
            QMessageBox.critical(self, "Hata", "Excel dosyası bulunamadı!")
        except pyodbc.Error as ex:
            QMessageBox.critical(self, "Hata", f"Veritabanı hatası oluştu!\n{ex}")
        except Exception as ex:
            QMessageBox.critical(self, "Hata", f"Bilinmeyen bir hata oluştu!\n{ex}")
        finally:
            conn.close()

    def SQLSorguyuCalistir(self):
        #BURAYA YAZ.

        #BUNU DENEMEYİ UNUTUMAAAAA
        query = self.SQLtextEdit.toPlainText().strip()
        if not query:
                return
        try:
            # Veritabanına bağlan
            conn = pyodbc.connect(conn_string)
            cursor = conn.cursor()
            
            # Saklı yordamı çağır
            cursor.execute(query)

            columns = [desc[0] for desc in cursor.description]
            self.SQLtableWidget.setColumnCount(len(columns))
            self.SQLtableWidget.setHorizontalHeaderLabels(columns)



            # Sonuçları al
            rows = cursor.fetchall()

            # TableWidget'i temizle
            self.SQLtableWidget.setRowCount(0)

            # Sonuçları TableWidget'e ekle
            for row_number, row_data in enumerate(rows):
                self.SQLtableWidget.insertRow(row_number)
                for column_number, data in enumerate(row_data):
                    item = QTableWidgetItem(str(data) if data is not None else "")
                    self.SQLtableWidget.setItem(row_number, column_number, item)

        except pyodbc.Error as ex:
            print("Veritabanı Hatası:", ex)
            QMessageBox.critical(self, "Hata", "Veritabanı bağlantı hatası!")

        finally:
            conn.close()

    def musteri_veri_cekme_fonksiyonu(self):

        try:
            # Veritabanına bağlan
            conn = pyodbc.connect(conn_string)
            cursor = conn.cursor()

            # Saklı yordamı çağır
            cursor.execute("SELECT * FROM Customer")

            rows = cursor.fetchall()
            
            # TableWidget'i temizle
            self.tableWidgetGuncelleSec.setRowCount(0)

            # Sonuçları TableWidget'e ekle
            for row_number, row_data in enumerate(rows):
                print(row_data)
                self.tableWidgetGuncelleSec.insertRow(row_number)
                for column_number, data in enumerate(row_data):
                    item = QTableWidgetItem(str(data) if data is not None else "")
                    self.tableWidgetGuncelleSec.setItem(row_number, column_number, item)

        except pyodbc.Error as ex:
            print("Veritabanı Hatası:", ex)
            QMessageBox.critical(self, "Hata", "Veritabanı bağlantı hatası!")
    
    def secili_satirdaki_verileri_linedite_cekme_fonksiyonu(self):
        """ Seçili satırdaki verileri LineEdit bileşenlerine aktarır """
        selected_row = self.tableWidgetGuncelleSec.currentRow()

        if selected_row != -1:
            # Seçili satırdaki verileri al ve LineEdit'lere yerleştir
            self.item0 = self.tableWidgetGuncelleSec.item(selected_row, 0) #id
            item1 = self.tableWidgetGuncelleSec.item(selected_row, 1)  # Ad (firstName)
            item2 = self.tableWidgetGuncelleSec.item(selected_row, 2)  # Soyad (lastName)
            item3 = self.tableWidgetGuncelleSec.item(selected_row, 3)  # Telefon (phone)
            item4 = self.tableWidgetGuncelleSec.item(selected_row, 4)  # Email (email)
            item5 = self.tableWidgetGuncelleSec.item(selected_row, 5)  # Şifre (Password)
            item6 = self.tableWidgetGuncelleSec.item(selected_row, 6)  # TC Kimlik No (tcKimlikNo)
            
            # Her item'i kontrol et ve LineEdit'e yerleştir
            if item1:
                self.line_edits0.setText(item1.text())  # Ad
            else:
                self.line_edits0.setText("")
                
            if item2:
                self.line_edits1.setText(item2.text())  # Soyad
            else:
                self.line_edits1.setText("")
                
            if item3:
                self.line_edits2.setText(item3.text())  # Telefon
            else:
                self.line_edits2.setText("")
                
            if item4:
                self.line_edits3.setText(item4.text())  # Email
            else:
                self.line_edits3.setText("")
                
            if item5:
                self.line_edits4.setText(item5.text())  # Şifre
            else:
                self.line_edits4.setText("")
                
            if item6:
                self.line_edits5.setText(item6.text())  # TC Kimlik No
            else:
                self.line_edits5.setText("")
            
            

    def lineeditteki_verileri_guncelleme_fonksiyonu(self):

        musAd=self.line_edits0.text()
        musSoyad=self.line_edits1.text()
        musTelNo=self.line_edits2.text()
        musMail=self.line_edits3.text()
        musSif=self.line_edits4.text()
        musKimlik=self.line_edits5.text()
        musSifTekrar=self.line_edits6.text()
        authorization=3
        
        if (musAd and musSoyad and musKimlik and musTelNo and musMail 
            and musSif and musSifTekrar and musSif == musSifTekrar):
            try:
                # Veritabanı bağlantısı
                conn = pyodbc.connect(conn_string)
                cursor = conn.cursor()

                # SQL sorgusu
                query = """
                UPDATE Customer
                SET firstName = ?, lastName = ?, phone = ?, email = ?, Password = ?, tcKimlikNo = ?
                WHERE customerID = ?
                """

                # Kullanıcı verilerini sorguya bağla ve çalıştır
                cursor.execute(query, (musAd, musSoyad, musTelNo, musMail, musSif, musKimlik, self.item0.text()))
                conn.commit()

                # Başarılı mesajı
                QMessageBox.information(self, "Başarılı", "Müşteri başarıyla güncellendi.")
            
            except pyodbc.Error as ex:
                # Hata mesajı
                print("Veritabanı Hatası:", ex)
                QMessageBox.critical(self, "Hata", "Müşteri güncellenirken bir hata oluştu.")
            
            finally:
                # Bağlantıyı kapatma
                conn.close()
        else:
            # Eksik bilgi veya şifre uyumsuzluğu uyarısı
            if musSif != musSifTekrar:
                QMessageBox.warning(self, "Uyarı", "Şifreler uyuşmuyor.")
            else:
                QMessageBox.warning(self, "Uyarı", "Lütfen tüm bilgileri doldurun.")
        
        self.item0.setText("-1")


class WindowMusAnaEkran(QMainWindow, Ui_MainWindowMusAnaEkran):
    def __init__(self):
        super().__init__()
        self.setupUi(self)
        #print(girisKimlikk.girisKimlikAd + girisKimlikk.girisKimlikSif)
        self.profilButton.clicked.connect(self.profil_buton)
        self.seferSorgulaButton.clicked.connect(self.load_data_per)
        self.biletAlButton.clicked.connect(self.bileti_al)



    def profil_buton(self):
        self.close()
        winMusProfil.show()


    def load_data_per(self):
        try:
            conn = pyodbc.connect(conn_string)
            cursor = conn.cursor()
            
            if self.kalkisNoktasi.text() and self.varisNoktasi.text():
                # Sorgunu çalıştır
                cursor.execute("SELECT * FROM JOURNEY WHERE departureCity='"+self.kalkisNoktasi.text()+"' AND arrivalCity='"+self.varisNoktasi.text()+"' AND departureDate='"+self.tarih.text()+"'")

                # Sonuçları bir listeye ata
                columns = [desc[0] for desc in cursor.description]
                self.tableWidgetSeferler.setColumnCount(len(columns) + 1)  # 1 ekstra sütun butonlar için
                self.tableWidgetSeferler.setHorizontalHeaderLabels(columns + ["Action"])

                # Sonuçları al
                rows = cursor.fetchall()

                # TableWidget'i temizle
                self.tableWidgetSeferler.setRowCount(0)

                # Sonuçları TableWidget'e ekle
                for row_number, row_data in enumerate(rows):
                    self.tableWidgetSeferler.insertRow(row_number)
                    
                    # Satır verilerini ekle
                    for column_number, data in enumerate(row_data):
                        item = QTableWidgetItem(str(data) if data is not None else "")
                        self.tableWidgetSeferler.setItem(row_number, column_number, item)

                    # Buton ekle (journeyID ile ilişkilendirilecek)
                    journey_id = row_data[0]  # Varsayalım ki journeyID birinci sütunda (index 0) yer alıyor
                    button = QPushButton(f"Seç")
                    button.clicked.connect(lambda checked, journey_id=journey_id: self.on_button_clicked(journey_id))
                    self.tableWidgetSeferler.setCellWidget(row_number, len(columns), button)

        except pyodbc.Error as ex:
            print(ex)

        finally:
            conn.close()
    
    def on_button_clicked(self,journey_id):
        self.journeyid2=journey_id
        QMessageBox.information(self, "Bilgi", f"{journey_id} Seçildi.")


    def bileti_al(self):
        insert_query = """
        INSERT INTO Ticket (seatNumber, purchaseDate, amount, customerID, journeyID)
        VALUES (?, ?, ?, ?, ?)
        """
        try:
            # Veritabanına bağlan
            conn = pyodbc.connect(conn_string)
            cursor = conn.cursor()
            
            # Saklı yordamı çağır
            cursor.execute("SELECT CustomerID FROM LogCustomer")
            for row in cursor.fetchall():
                self.customerID_2 = str(row[0]).strip()
            results=cursor.execute("SELECT departureDate, amount FROM Journey WHERE journeyID = ?", self.journeyid2)
            if results:
                for row in results:
                    purchaseDate = row[0]
                    amount = row[1]

            cursor.execute(insert_query, (self.comboBoxKoltuk.currentText(), purchaseDate, amount, self.customerID_2, self.journeyid2))
            conn.commit()
            QMessageBox.information(self, "Hata", "Başarılı")

        except pyodbc.Error as ex:
            print("Veritabanı Hatası:", ex)
            QMessageBox.critical(self, "Hata", "Veritabanı bağlantı hatası!")

        finally:
            conn.close()
        


    def widgete_getir(self):
        try:
            # Veritabanına bağlan
            conn = pyodbc.connect(conn_string)
            cursor = conn.cursor()
            
            # Saklı yordamı çağır
            cursor.execute("SELECT * FROM Ticket")

            columns = [desc[0] for desc in cursor.description]
            self.tableWidgetSeferler.setColumnCount(len(columns))
            self.tableWidgetSeferler.setHorizontalHeaderLabels(columns)



            # Sonuçları al
            rows = cursor.fetchall()

            # TableWidget'i temizle
            self.tableWidgetSeferler.setRowCount(0)

            # Sonuçları TableWidget'e ekle
            for row_number, row_data in enumerate(rows):
                self.tableWidgetSeferler.insertRow(row_number)
                for column_number, data in enumerate(row_data):
                    item = QTableWidgetItem(str(data) if data is not None else "")
                    self.tableWidgetSeferler.setItem(row_number, column_number, item)

        except pyodbc.Error as ex:
            print("Veritabanı Hatası:", ex)
            QMessageBox.critical(self, "Hata", "Veritabanı bağlantı hatası!")

        finally:
            conn.close()



class WindowPerProfilEkrani(QMainWindow, Ui_MainWindowPerProfil):
    def __init__(self):
        super().__init__()
        self.setupUi(self)
        self.yenile
        self.anaSayfaButton.clicked.connect(self.anasayfa)
        self.yenileButton.clicked.connect(self.yenile)
        #self.perGirisButton.clicked.connect(self.load_data_per)

        


    def anasayfa(self):
        winPerAnaEkran.show()
        self.close()

    def yenile(self):
        try:
            conn = pyodbc.connect(conn_string)
            cursor = conn.cursor()

            cursor.execute("SELECT employeeID, firstName, lastName, department, phone, username FROM LogEmployee")
            for row in cursor.fetchall():
                self.adSoyadLabel.setText(f"{row.firstName} {row.lastName}")
                self.departmanLabel.setText(f"{row.department}")
                self.telNoLabel.setText(f"{row.phone}")
                self.usernameLabel.setText(f"{row.username}")


        except pyodbc.Error as ex:
            print(ex)

        finally:
            conn.close()

class WindowMusProfilEkrani(QMainWindow, Ui_MainWindowMusProfil):
    def __init__(self):
        super().__init__()
        self.setupUi(self)
        self.yenile()
        self.anaSayfaButton.clicked.connect(self.anasayfa)
        self.yenileButton.clicked.connect(self.yenile)
        #self.perGirisButton.clicked.connect(self.load_data_per)

    def anasayfa(self):
        winMusAnaEkran.show()
        self.close()


    def yenile(self):
        try:
            conn = pyodbc.connect(conn_string)
            cursor = conn.cursor()

            cursor.execute("SELECT CustomerID, firstName, lastName, phone, email, password, tcKimlikNo, authorizationID FROM LogCustomer")
            for row in cursor.fetchall():
                self.adSoyadLabel.setText(f"{row.firstName} {row.lastName}")
                self.kimlikNoLabel.setText(f"{row.tcKimlikNo}")
                self.telNoLabel.setText(f"{row.phone}")
                self.mailLabel.setText(f"{row.email}")

        except pyodbc.Error as ex:
            print(ex)

        finally:
            conn.close()


class SqlBackupApp(QWidget):
    def __init__(self):
        super().__init__()

        self.setWindowTitle('SQL Server Yedekleme')
        self.setGeometry(200, 200, 500, 250)

        # UI bileşenlerini başlat
        self.init_ui()

    def init_ui(self):
        layout = QVBoxLayout()

        # SQL Server bilgisi
        self.db_server_label = QLabel('SQL Server Adresi (localhost):', self)
        layout.addWidget(self.db_server_label)

        self.db_server_input = QLineEdit(self)
        self.db_server_input.setText('localhost')  # Varsayılan olarak localhost
        layout.addWidget(self.db_server_input)

        self.db_name_label = QLabel('Veritabanı Adı: ', self)
        layout.addWidget(self.db_name_label)

        self.db_name_input = QLineEdit(self)  # Veritabanı adını girilecek alan
        self.db_name_input.setText('journey_management')  # Varsayılan olarak 'journey_management'
        layout.addWidget(self.db_name_input)

        # Yedekleme butonu
        self.backup_button = QPushButton('Veritabanını Yedekle', self)
        self.ana_sayfa_button = QPushButton('Ana Sayfa', self)
        self.backup_button.clicked.connect(self.start_backup)
        self.ana_sayfa_button.clicked.connect(self.anasayfa)
        layout.addWidget(self.backup_button)
        layout.addWidget(self.ana_sayfa_button)

        # Çıktı mesajı
        self.output_label = QLabel('Çıktı:', self)
        layout.addWidget(self.output_label)

        self.output_text = QLabel('', self)
        layout.addWidget(self.output_text)

        # Layout'u ayarla
        self.setLayout(layout)

    def anasayfa(self):
        winPerAnaEkran.show()
        self.close()



    def start_backup(self):
        # Arka planda yedekleme işlemi başlat
        self.backup_button.setEnabled(False)  # Butonu devre dışı bırak
        self.output_text.setText("Yedekleme işlemi başlatılıyor...")
        QApplication.processEvents()  # GUI'nin donmaması için bu çağrıyı ekliyoruz

        # Arka planda yedekleme işlemi çalıştırılıyor
        server = self.db_server_input.text()
        database = self.db_name_input.text()

        # Yedekleme dosyasını seçme
        backup_dir = "C:\\BACKUP_DB"
        if not os.path.exists(backup_dir):
            os.makedirs(backup_dir)

        file_name = os.path.join(backup_dir, f"{database}FullBackup{self.get_current_timestamp()}.bak")

        if file_name:
            self.backup_database(server, database, file_name)
        else:
            self.output_text.setText("Yedekleme iptal edildi.")
            self.backup_button.setEnabled(True)

    def get_current_timestamp(self):
        # Mevcut tarihi ve saati formatlayarak yedek dosyasının adını oluşturur
        from datetime import datetime
        return datetime.now().strftime("%Y-%m-%d_%H-%M-%S")

    def backup_database(self, server, database, file_name):
        """
        Veritabanını belirtilen dizine yedekler.
        """

        # Windows Authentication kullanarak SQL Server'a bağlan
        conn_str = f'DRIVER={{ODBC Driver 17 for SQL Server}};SERVER={server};DATABASE=master;Trusted_Connection=yes;'

        try:
            # SQL Server'a bağlan
            cnxn = pyodbc.connect(conn_str)
            cnxn.autocommit = True  # Komutların otomatik olarak commit edilmesi
            cursor = cnxn.cursor()

            # Veritabanını SINGLE_USER moduna al
            cursor.execute(f"ALTER DATABASE [{database}] SET SINGLE_USER WITH ROLLBACK IMMEDIATE;")
            cursor.commit()  # Değişiklikleri kaydet

            # Yedekleme SQL komutunu oluştur
            backup_cmd = f"""
            BACKUP DATABASE [{database}]
            TO DISK = N'{file_name}'
            WITH NOFORMAT, NOINIT, NAME = N'{database}-Full Database Backup', SKIP, NOREWIND, NOUNLOAD, STATS = 10 ;
            """
            cursor.execute(backup_cmd)  # Yedekleme komutunu çalıştır
            while (cursor.nextset()):
                pass
            cursor.commit()  # Değişiklikleri kaydet

            # Yedekleme tamamlandığında mesaj göster
            self.output_text.setText(f"Yedekleme başarıyla tamamlandı: {file_name}")

            # Veritabanını MULTI_USER moduna al
            cursor.execute(f"ALTER DATABASE [{database}] SET MULTI_USER;")
            cursor.commit()  # Değişiklikleri kaydet

            # Dosyanın kaydedilip kaydedilmediğini kontrol et
            if os.path.exists(file_name):
                self.output_text.setText(f"Yedekleme dosyası başarıyla kaydedildi: {file_name}")
            else:
                self.output_text.setText(f"Yedekleme dosyası kaydedilemedi: {file_name}")

        except Exception as e:
            # Hata mesajı
            self.output_text.setText(f"Yedekleme hatası: {e}")
            print(f"Hata: {e}")
        finally:
            if 'cursor' in locals() and cursor:
                cursor.close()
            if 'cnxn' in locals() and cnxn:
                cnxn.close()
            self.backup_button.setEnabled(True)  # Yedekleme tamamlandığında butonu tekrar etkinleştir

class BackupRestoreApp(QWidget):
    def __init__(self):
        super().__init__()

        self.setWindowTitle('SQL Server Geri Yükleme Uygulaması')
        self.setGeometry(100, 100, 600, 400)

        # Layout ve bileşenler
        self.layout = QVBoxLayout()

        self.label = QLabel('Geri Yüklemek için Yedek Dosyasını Seçin:', self)
        self.layout.addWidget(self.label)

        self.select_button = QPushButton('Yedek Dosyasını Seç', self)
        self.anasayfa_button = QPushButton('Ana Sayfa', self)
        self.select_button.clicked.connect(self.select_backup_file)
        self.anasayfa_button.clicked.connect(self.anasayfa)
        self.layout.addWidget(self.select_button)
        self.layout.addWidget(self.anasayfa_button)

        # Veritabanı seçim kutusu (ComboBox)
        self.db_selector_label = QLabel('Geri Yüklenecek Veritabanını Seçin:', self)
        self.layout.addWidget(self.db_selector_label)

        self.db_selector = QComboBox(self)
        self.layout.addWidget(self.db_selector)

        self.restore_button = QPushButton('Geri Yükle', self)
        self.restore_button.clicked.connect(self.restore_database)
        self.layout.addWidget(self.restore_button)

        # İşlem çıktısını gösterecek TextEdit widget'ı
        self.output_box = QTextEdit(self)
        self.output_box.setReadOnly(True)
        self.layout.addWidget(self.output_box)

        self.setLayout(self.layout)

        # Bağlantı parametreleri (localhost üzerinden bağlanma)
        self.conn_str = (
            r'DRIVER={ODBC Driver 17 for SQL Server};'
            r'SERVER=localhost;'  # Buraya localhost ya da 127.0.0.1 yazabilirsiniz
            r'DATABASE=master;'  # Bu, master veritabanına bağlanır
            r'Trusted_Connection=yes;'  # Windows Authentication
        )
        self.backup_file = None  # Yedek dosyasının yolu

        # Veritabanlarını yükle
        self.load_databases()
    
    def anasayfa(self):
        winPerAnaEkran.show()
        self.close()


    def load_databases(self):
        """SQL Server'daki veritabanlarını yükleyin ve combobox'a ekleyin."""
        try:
            # SQL Server'a bağlan
            conn = pyodbc.connect(self.conn_str)
            cursor = conn.cursor()

            # Mevcut veritabanlarını sorgula
            cursor.execute("SELECT name FROM sys.databases WHERE state_desc = 'ONLINE' AND name NOT IN ('master', 'tempdb', 'model', 'msdb')")
            databases = cursor.fetchall()

            # Veritabanlarını combobox'a ekle
            for db in databases:
                self.db_selector.addItem(db[0])

            cursor.close()
            conn.close()

        except Exception as e:
            self.output_box.append(f"Veritabanları yüklenirken bir hata oluştu: {str(e)}")

    def select_backup_file(self):
        """Kullanıcıdan yedek dosyasını seçmesini isteyin."""
        options = QFileDialog.Options()
        self.backup_file, _ = QFileDialog.getOpenFileName(self, "Yedek Dosyasını Seç", "C:\\BACKUP_DB", "Backup Files (*.bak);;All Files (*)")
        if self.backup_file:
            self.label.setText(f"Seçilen Yedek Dosyası: {self.backup_file}")
        else:
            self.label.setText("Yedek dosyası seçilmedi!")

    def restore_database(self):
        """Veritabanını yedek dosyasından geri yükleyin ve ilerlemeyi gösterin."""
        db_name = self.db_selector.currentText()
        if not db_name:
            self.label.setText("Lütfen geçerli bir veritabanı seçin.")
            return

        if not self.backup_file:
            self.label.setText("Lütfen geçerli bir yedek dosyası seçin.")
            return

        try:
            # SQL Server'a bağlan ve autocommit=True yap
            conn = pyodbc.connect(self.conn_str)
            conn.autocommit = True
            cursor = conn.cursor()

            # Geri yükleme komutunu oluştur
            restore_cmd = f"""
            RESTORE DATABASE [{db_name}]
            FROM DISK = N'{self.backup_file}'
            WITH REPLACE,
                 FILE = 1,
                 STATS = 1;  -- Bu komut, her %1'lik ilerlemeyi rapor eder
            """

            # Komut çalıştır
            cursor.execute(restore_cmd)

            # Sonuç kümesini (result set) kontrol et ve işlem ilerlemelerini al
            while cursor.nextset():  # Sonraki küme varsa devam et
                pass

            # Geri yükleme işlemi tamamlandıktan sonra veritabanını MULTI_USER moduna al
            cursor.execute(f"ALTER DATABASE [{db_name}] SET MULTI_USER;")

            # Sonuç mesajlarını ekrana yazdır
            self.output_box.append("Geri yükleme işlemi tamamlandı ve veritabanı MULTI_USER moduna alındı.")
            self.label.setText(f"{db_name} veritabanı başarıyla geri yüklendi ve MULTI_USER moduna alındı.")

            # Bağlantıyı kapat
            cursor.close()
            conn.close()

        except Exception as e:
            # Hata mesajını ekrana yazdır
            self.label.setText(f"Bir hata oluştu: {str(e)}")
            self.output_box.append(f"Error: {str(e)}")



if __name__ == "__main__":
    app = QApplication(sys.argv)
    winGiris = WindowGiris()
    #winBaglan = WindowBaglan()
    winPerAnaEkran = WindowPerAnaEkran()
    winMusAnaEkran = WindowMusAnaEkran()
    winPerProfil = WindowPerProfilEkrani()
    winMusProfil = WindowMusProfilEkrani()
    winBackup = SqlBackupApp()
    winBackupRestore = BackupRestoreApp()
    #winBaglan.show()
    winGiris.show()

    sys.exit(app.exec())
