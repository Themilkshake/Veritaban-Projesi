# python C:\Users\lenovo\Desktop\MEKATEK_MUY\pyqt6_yer_istasyonu\yer_istasyonu_3.py & python C:\Users\lenovo\Desktop\MEKATEK_MUY\pyqt6_yer_istasyonu\video_html\app.py
import sys
from openpyxl import Workbook, load_workbook
from PyQt5.QtWidgets import QApplication,QMainWindow,QMessageBox,QTableWidgetItem,QFileDialog
from personel_ana_ekrani_UI import Ui_MainWindowPerAnaEkran
import pyodbc


#------------------------------------------------
# Veritabanı bağlantı bilgilerini buraya girin
conn_string = (
    r'DRIVER={SQL Server};'
    r'SERVER=ALI;'
    r'DATABASE=journey_management2;'
    r'UID=;'
    r'PWD=;'
)
#--------------------------------------------------

class Window(QMainWindow, Ui_MainWindowPerAnaEkran):
    def __init__(self):
        super().__init__()
        self.setupUi(self)

        self.yenile()
        self.cikisButton.clicked.connect(self.cikis)
        self.yenileButton.clicked.connect(self.yenile)
        self.profilButton.clicked.connect(self.profil)

        self.SeferTanimlaButton.clicked.connect(self.seferKontrol)
        self.aracTanimlaButton.clicked.connect(self.aracKontrol)
        self.musTanimlaButton.clicked.connect(self.musteri_ekleme)
        self.surucuTanimlaButton.clicked.connect(self.surucu_ekleme)
        
        
        self.seferiIptalEtButton.clicked.connect(self.sefer_silme)
        self.musSilButton.clicked.connect(self.musteri_silme)
        self.surucuSilButton.clicked.connect(self.surucu_silme)

        self.seferSorgulamafiltreButton.clicked.connect(self.sefer_sorgulama)
        self.aracSorgulamafiltreButton.clicked.connect(self.arac_sorgulama)
        self.musSorgulamafiltreButton.clicked.connect(self.musteri_sorgulama)
        self.surucuSorgulamafiltreButton.clicked.connect(self.surucu_sorgulama)

        self.excelSeferButton.clicked.connect(self.sefer_filtreleme)
        self.excelAracButton.clicked.connect(self.arac_filtreleme)
        self.excelMusteriButton.clicked.connect(self.musteri_filtreleme)
        self.excelSurucuButton.clicked.connect(self.surucu_filtreleme)
        self.exceldenVeritabaninaKaydetButton.clicked.connect(self.driver_excel_veri_aktar)


    def profil(self):
        self.close()


    def yenile(self):
        self.comboboxSilinecekSefer.clear()
        self.comboboxSilinecekArac.clear()
        self.comboboxSilinecekMusteri.clear()
        self.comboboxSilinecekSurucu.clear()

        self.comboboxSilinecekSefer.addItem(" ")
        self.comboboxSilinecekArac.addItem(" ")
        self.comboboxSilinecekMusteri.addItem(" ")
        self.comboboxSilinecekSurucu.addItem(" ")

        self.comboboxKalkisSaati.clear()
        self.comboboxVarisSaati.clear()
        self.comboboxKalkisYeriFiltre.clear()
        self.comboboxVarisYeriFiltre.clear()
        self.comboboxAracMarka_2.clear()
        self.comboboxAracModel_2.clear()
        self.comboboxAracPlaka.clear()

        self.comboboxKalkisSaati.addItem(" ")
        self.comboboxVarisSaati.addItem(" ")
        self.comboboxKalkisYeriFiltre.addItem(" ") 
        self.comboboxVarisYeriFiltre.addItem(" ")
        self.comboboxAracMarka_2.addItem(" ") 
        self.comboboxAracModel_2.addItem(" ") 
        self.comboboxAracPlaka.addItem(" ") 


        self.aracSurucuGetirSql()
        self.aracOzellikGetirSql()
        self.SeferveAracSrogulamaFiltrelemeGetir()
        self.silinecek_arac_musteri_surucu_getir_sql()

    def aracSurucuGetirSql(self):
        try:
            conn = pyodbc.connect(conn_string)
            cursor = conn.cursor()

            cursor.execute("SELECT driverID, firstName, lastName FROM Driver")
            for row in cursor.fetchall():
                self.comboboxSurucuSecimi.addItem(f"{row.driverID} {row.firstName} {row.lastName}")
            
            cursor.execute("SELECT busID, marka FROM Bus")
            for row in cursor.fetchall():
                self.comboboxAracSecimi.addItem(f"{row.busID} {row.marka}")    
        except pyodbc.Error as ex:
            print(ex)

        finally:
            conn.close()

    def aracOzellikGetirSql(self):
        try:
            conn = pyodbc.connect(conn_string)
            cursor = conn.cursor()

            cursor.execute("SELECT bustypeID, Feature FROM BusType")
            for row in cursor.fetchall():
                self.comboboxAracOzellikleri.addItem(f"{row.bustypeID} {row.Feature}")   
        except pyodbc.Error as ex:
            print(ex)

        finally:
            conn.close()

    def SeferveAracSrogulamaFiltrelemeGetir(self):
        try:
            conn = pyodbc.connect(conn_string)
            cursor = conn.cursor()

            cursor.execute("SELECT departureDate, departureCity, arrivalDate, arrivalCity FROM Journey")
            for row in cursor.fetchall():
                self.comboboxKalkisSaati.addItem(f"{row.departureDate}")
                self.comboboxVarisSaati.addItem(f"{row.arrivalDate}")
                self.comboboxKalkisYeriFiltre.addItem(f"{row.departureCity}")
                self.comboboxVarisYeriFiltre.addItem(f"{row.arrivalCity}") 
            
            cursor.execute("SELECT marka, model, numberPlate FROM Bus")
            for row in cursor.fetchall():
                self.comboboxAracMarka_2.addItem(f"{row.marka}") 
                self.comboboxAracModel_2.addItem(f"{row.model}") 
                self.comboboxAracPlaka.addItem(f"{row.numberPlate}") 


        except pyodbc.Error as ex:
            print(ex)

        finally:
            conn.close()


    def cikis(self):
        reply = QMessageBox.question(self, '!', 'Çıkmak istediğinize emin misiniz?', QMessageBox.Yes | QMessageBox.No)
        if reply == QMessageBox.Yes:
            app.quit()

    def seferKontrol(self):
        reply = QMessageBox.question(self, '!', 'Seferi kaydetmek istadiğinizden emin misiniz?', QMessageBox.Yes | QMessageBox.No)
        if reply == QMessageBox.Yes:
            self.sefer_tanimla()
        
    def sefer_tanimla(self):
        kalkisNoktasi=self.kalkis_noktasi.text()
        varisNoktasi=self.varis_noktasi.text()
        kalkisTarih=self.kalkis_tarih.text()
        varisTarih=self.aracModel.text()
        #sürücü ve araç seçimi kodu burada olacak.

        AracSecimi = self.comboboxAracSecimi.currentText()
        SurucuSecimi = self.comboboxSurucuSecimi.currentText()
        driverID = SurucuSecimi.split(" ", 1)[0]  # Boşluktan ayırma
        busID= AracSecimi.split(" ", 1)[0]

        if kalkisNoktasi and varisNoktasi:
            try:
                conn = pyodbc.connect(conn_string)
                cursor = conn.cursor()
                query = """
                INSERT INTO Journey (departureCity, arrivalCity, departureDate, arrivalDate, busID, driverID)
                VALUES (?, ?, ?, ?, ?, ?)
                """
                # Kullanıcı verilerini sorguya bağla ve çalıştır
                cursor.execute(query, (kalkisNoktasi, varisNoktasi, kalkisTarih, varisTarih, busID, driverID))
                conn.commit()

            except pyodbc.Error as ex:
                print(ex)

            finally:
                conn.close()
        else:
            QMessageBox.warning(self, "Uyarı", "Kalkış Noktası ve Varış Noktası giriniz.")

        #Buton ayarlanacak.

    def aracKontrol(self):
        reply = QMessageBox.question(self, '!', 'Aracı kaydetmek istadiğinizden emin misiniz?', QMessageBox.Yes | QMessageBox.No)
        if reply == QMessageBox.Yes:
            self.arac_tanimla()
    
    def arac_tanimla(self):
        aracKapasite=self.aracKapasite.text()
        aracPlakaKodu=self.aracPlakaKodu.text()
        aracMarka=self.aracMarka.text()
        aracModel=self.aracModel.text()
        AracOzellikleri = self.comboboxAracOzellikleri.currentText()
        bustypeID, Feature = AracOzellikleri.split(" ",1)  # Boşluktan ayırma

        if aracKapasite and aracPlakaKodu and aracMarka and aracModel and bustypeID:
            try:
                # Veritabanı bağlantısı
                conn = pyodbc.connect(conn_string)
                cursor = conn.cursor()

                # SQL sorgusu
                query = """
                INSERT INTO Bus (capacity, numberPlate, marka, bustypeID, model)
                VALUES (?, ?, ?, ?, ?)
                """

                # Kullanıcı verilerini sorguya bağla ve çalıştır
                cursor.execute(query, (aracKapasite, aracPlakaKodu, aracMarka, bustypeID, aracModel))
                conn.commit()
                
                # Başarılı mesajı
                QMessageBox.information(self, "Başarılı", "Araç başarıyla eklendi.")
            
            except pyodbc.Error as ex:
                # Hata mesajı
                print("Veritabanı Hatası:", ex)
                QMessageBox.critical(self, "Hata", "Araç eklenirken bir hata oluştu.")
            
            finally:
                # Bağlantıyı kapatma
                conn.close()
        else:
            # Eksik bilgi uyarısı
            QMessageBox.warning(self, "Uyarı", "Lütfen tüm bilgileri doldurun.")
    
    def musteri_ekleme(self):
        musAd=self.musAd.text()
        musSoyad=self.musSoyad.text()
        musKimlik=self.musKimlik.text()
        musTelNo=self.musTelNo.text()
        musMail=self.musMail.text()
        musSif=self.musSif.text()
        musSifTekrar=self.musSifTekrar.text()
        authorization=3


        if (musAd and musSoyad and musKimlik and musTelNo and musMail 
            and musSif and musSifTekrar and musSif == musSifTekrar):
            try:
                # Veritabanı bağlantısı
                conn = pyodbc.connect(conn_string)
                cursor = conn.cursor()

                # SQL sorgusu
                query = """
                INSERT INTO Customer (firstName, lastName, phone, email, Password, tcKimlikNo, Authorization)
                VALUES (?, ?, ?, ?, ?, ?, ?)
                """

                # Kullanıcı verilerini sorguya bağla ve çalıştır
                cursor.execute(query, (musAd, musSoyad, musTelNo, musMail, musSif, musKimlik, authorization))
                conn.commit()

                # Başarılı mesajı
                QMessageBox.information(self, "Başarılı", "Müşteri başarıyla eklendi.")
            
            except pyodbc.Error as ex:
                # Hata mesajı
                print("Veritabanı Hatası:", ex)
                QMessageBox.critical(self, "Hata", "Müşteri eklenirken bir hata oluştu.")
            
            finally:
                # Bağlantıyı kapatma
                conn.close()
        else:
            # Eksik bilgi veya şifre uyumsuzluğu uyarısı
            if musSif != musSifTekrar:
                QMessageBox.warning(self, "Uyarı", "Şifreler uyuşmuyor.")
            else:
                QMessageBox.warning(self, "Uyarı", "Lütfen tüm bilgileri doldurun.")

    def surucu_ekleme(self):
        # Kullanıcıdan alınacak veriler
        driverFirstName = self.surucuAd.text()  # QLineEdit'den ad bilgisi
        driverLastName = self.surucuSoyad.text()  # QLineEdit'den soyad bilgisi
        driverPhone = self.surucuTelNo.text()  # QLineEdit'den telefon numarası
        driverLicenseNumber = self.surucuLisansNo.text()  # QLineEdit'den ehliyet numarası

        # Boş alan kontrolü
        if driverFirstName and driverLastName and driverPhone and driverLicenseNumber:
            try:
                # Veritabanı bağlantısı
                conn = pyodbc.connect(conn_string)
                cursor = conn.cursor()

                # SQL sorgusu
                query = """
                INSERT INTO Driver (firstName, lastName, phone, licenseNumber)
                VALUES (?, ?, ?, ?)
                """

                # Kullanıcı verilerini sorguya bağla ve çalıştır
                cursor.execute(query, (driverFirstName, driverLastName, driverPhone, driverLicenseNumber))
                conn.commit()

                # Başarılı mesajı
                QMessageBox.information(self, "Başarılı", "Sürücü başarıyla eklendi.")
            
            except pyodbc.Error as ex:
                # Hata mesajı
                print("Veritabanı Hatası:", ex)
                QMessageBox.critical(self, "Hata", "Sürücü eklenirken bir hata oluştu.")
            
            finally:
                # Bağlantıyı kapatma
                conn.close()
        else:
            # Eksik bilgi uyarısı
            QMessageBox.warning(self, "Uyarı", "Lütfen tüm bilgileri doldurun.")

    def sefer_sorgulama(self):
        departureDatee = self.comboboxKalkisSaati.currentText()
        arrivalDatee = self.comboboxVarisSaati.currentText()
        departureCityy = self.comboboxKalkisYeriFiltre.currentText()
        arrivalCityy = self.comboboxVarisYeriFiltre.currentText()
        
        query = "SELECT * FROM Journey WHERE 1=1"
        params = []

        # Filtreleri ekle
        if departureCityy != " ":
            query += " AND departureCity = ?"
            params.append(departureCityy)

        if arrivalCityy != " ":
            query += " AND arrivalCity = ?"
            params.append(arrivalCityy)

        if departureDatee != " ":
            query += " AND departureDate = ?"
            params.append(departureDatee)

        if arrivalDatee != " ":
            query += " AND arrivalDate = ?"
            params.append(arrivalDatee)
        
        print(params)
        try:
            # Veritabanına bağlan ve sorguyu çalıştır
            conn = pyodbc.connect(conn_string)
            cursor = conn.cursor()
            cursor.execute(query,params)
            rows = cursor.fetchall()
            # TableWidget'i temizle
            self.tableWidgetSeferler.setRowCount(0)

            # Sorgu sonuçlarını TableWidget'e ekle
            for row_number, row_data in enumerate(rows):
                
                self.tableWidgetSeferler.insertRow(row_number)
                for column_number, data in enumerate(row_data):

                    item = QTableWidgetItem(str(data) if data is not None else "") 
                    self.tableWidgetSeferler.setItem(row_number, column_number, item)
        except pyodbc.Error as ex:
            print("Veritabanı Hatası:", ex)
            QMessageBox.critical(self, "Hata", "Veritabanı bağlantı hatası!")

        finally:
            conn.close()

    def sefer_filtreleme(self):

        dosya_yolu, _ = QFileDialog.getSaveFileName(
            self, "Excel Dosyasını Kaydet", "", "Excel Dosyası (*.xlsx)"
        )
        
        if not dosya_yolu:
            return  # Dosya seçilmedi

        try:
            # Excel çalışma kitabı oluştur
            wb = Workbook()
            ws = wb.active
            ws.title = "Seferler"

            # TableWidget başlıklarını yaz
            kolon_sayisi = self.tableWidgetSeferler.columnCount()
            satir_sayisi = self.tableWidgetSeferler.rowCount()

            # Başlıkları Excel'e ekle
            for kolon in range(kolon_sayisi):
                baslik = self.tableWidgetSeferler.horizontalHeaderItem(kolon).text()
                ws.cell(row=1, column=kolon + 1, value=baslik)

            # TableWidget verilerini Excel'e ekle
            for satir in range(satir_sayisi):
                for kolon in range(kolon_sayisi):
                    hucre_verisi = self.tableWidgetSeferler.item(satir, kolon)
                    ws.cell(row=satir + 2, column=kolon + 1, value=hucre_verisi.text() if hucre_verisi else "")

            # Dosyayı kaydet
            wb.save(dosya_yolu)
            QMessageBox.information(self, "Başarılı", "Veriler Excel dosyasına başarıyla aktarıldı!")
        
        except Exception as ex:
            QMessageBox.critical(self, "Hata", f"Dosya kaydedilirken hata oluştu!\n{ex}")


    def arac_sorgulama(self):
        plaka_2 = self.comboboxAracPlaka.currentText()
        marka_2 = self.comboboxAracMarka_2.currentText()
        model_2 = self.comboboxAracModel_2.currentText()
        
        query = "SELECT busID, capacity, numberPlate, marka, model FROM Bus WHERE 1=1"
        params = []

        # Filtreleri ekle
        if plaka_2 != " ":
            query += " AND numberPlate = ?"
            params.append(plaka_2)

        if marka_2 != " ":
            query += " AND marka = ?"
            params.append(marka_2)

        if model_2 != " ":
            query += " AND model = ?"
            params.append(model_2)

        
        try:
            # Veritabanına bağlan ve sorguyu çalıştır
            conn = pyodbc.connect(conn_string)
            cursor = conn.cursor()
            cursor.execute(query,params)
            rows = cursor.fetchall()
            # TableWidget'i temizle
            self.aracSorgulama_tableWidget.setRowCount(0)

            # Sorgu sonuçlarını TableWidget'e ekle
            for row_number, row_data in enumerate(rows):
                
                self.aracSorgulama_tableWidget.insertRow(row_number)
                for column_number, data in enumerate(row_data):

                    item = QTableWidgetItem(str(data) if data is not None else "") 
                    self.aracSorgulama_tableWidget.setItem(row_number, column_number, item)
        except pyodbc.Error as ex:
            print("Veritabanı Hatası:", ex)
            QMessageBox.critical(self, "Hata", "Veritabanı bağlantı hatası!")

        finally:
            conn.close()

    def arac_filtreleme(self):

        dosya_yolu, _ = QFileDialog.getSaveFileName(
            self, "Excel Dosyasını Kaydet", "", "Excel Dosyası (*.xlsx)"
        )
        
        if not dosya_yolu:
            return  # Dosya seçilmedi

        try:
            # Excel çalışma kitabı oluştur
            wb = Workbook()
            ws = wb.active
            ws.title = "Seferler"

            # TableWidget başlıklarını yaz
            kolon_sayisi = self.aracSorgulama_tableWidget.columnCount()
            satir_sayisi = self.aracSorgulama_tableWidget.rowCount()

            # Başlıkları Excel'e ekle
            for kolon in range(kolon_sayisi):
                baslik = self.aracSorgulama_tableWidget.horizontalHeaderItem(kolon).text()
                ws.cell(row=1, column=kolon + 1, value=baslik)

            # TableWidget verilerini Excel'e ekle
            for satir in range(satir_sayisi):
                for kolon in range(kolon_sayisi):
                    hucre_verisi = self.aracSorgulama_tableWidget.item(satir, kolon)
                    ws.cell(row=satir + 2, column=kolon + 1, value=hucre_verisi.text() if hucre_verisi else "")

            # Dosyayı kaydet
            wb.save(dosya_yolu)
            QMessageBox.information(self, "Başarılı", "Veriler Excel dosyasına başarıyla aktarıldı!")
        
        except Exception as ex:
            QMessageBox.critical(self, "Hata", f"Dosya kaydedilirken hata oluştu!\n{ex}")
    
    def musteri_sorgulama(self):

        # ComboBox değerlerini al
        firstName = self.musSorgulama_MusAd.text() or None
        lastName = self.musSorgulama_MusSoyad.text() or None
        phone = self.musSorgulama_MusTelNo.text() or None
        email = self.musSorgulama_MusMail.text() or None

        try:
            # Veritabanına bağlan
            conn = pyodbc.connect(conn_string)
            cursor = conn.cursor()

            # Saklı yordamı çağır
            cursor.execute("""
                EXEC SearchCustomers 
                    @firstName = ?, 
                    @lastName = ?, 
                    @phone = ?, 
                    @email = ?
            """, firstName, lastName, phone, email)

            rows = cursor.fetchall()
            
            # TableWidget'i temizle
            self.musSorgulama_tableWidget.setRowCount(0)

            # Sonuçları TableWidget'e ekle
            for row_number, row_data in enumerate(rows):
                print(row_data)
                self.musSorgulama_tableWidget.insertRow(row_number)
                for column_number, data in enumerate(row_data):
                    item = QTableWidgetItem(str(data) if data is not None else "")
                    self.musSorgulama_tableWidget.setItem(row_number, column_number, item)

        except pyodbc.Error as ex:
            print("Veritabanı Hatası:", ex)
            QMessageBox.critical(self, "Hata", "Veritabanı bağlantı hatası!")

        finally:
            conn.close()

    def musteri_filtreleme(self):

        dosya_yolu, _ = QFileDialog.getSaveFileName(
            self, "Excel Dosyasını Kaydet", "", "Excel Dosyası (*.xlsx)"
        )
        
        if not dosya_yolu:
            return  # Dosya seçilmedi

        try:
            # Excel çalışma kitabı oluştur
            wb = Workbook()
            ws = wb.active
            ws.title = "Seferler"

            # TableWidget başlıklarını yaz
            kolon_sayisi = self.musSorgulama_tableWidget.columnCount()
            satir_sayisi = self.musSorgulama_tableWidget.rowCount()

            # Başlıkları Excel'e ekle
            for kolon in range(kolon_sayisi):
                baslik = self.musSorgulama_tableWidget.horizontalHeaderItem(kolon).text()
                ws.cell(row=1, column=kolon + 1, value=baslik)

            # TableWidget verilerini Excel'e ekle
            for satir in range(satir_sayisi):
                for kolon in range(kolon_sayisi):
                    hucre_verisi = self.musSorgulama_tableWidget.item(satir, kolon)
                    ws.cell(row=satir + 2, column=kolon + 1, value=hucre_verisi.text() if hucre_verisi else "")

            # Dosyayı kaydet
            wb.save(dosya_yolu)
            QMessageBox.information(self, "Başarılı", "Veriler Excel dosyasına başarıyla aktarıldı!")
        
        except Exception as ex:
            QMessageBox.critical(self, "Hata", f"Dosya kaydedilirken hata oluştu!\n{ex}")

    def surucu_sorgulama(self):

        # ComboBox değerlerini al
        surucuAd = self.surucuSorgulama_SurucuAd.text() or None
        surucuSoyad = self.surucuSorgulama_SurucuSoyad.text() or None
        surucutelNo = self.surucuSorgulama_SurucuTelno.text() or None
        suruculinsansNo = self.surucuSorgulama_SurucuLisansNo.text() or None

        try:
            # Veritabanına bağlan
            conn = pyodbc.connect(conn_string)
            cursor = conn.cursor()

            # Saklı yordamı çağır
            cursor.execute("""
                EXEC SearchDrivers 
                    @firstName = ?, 
                    @lastName = ?, 
                    @phone = ?, 
                    @licenseNumber = ?
            """, surucuAd, surucuSoyad, surucutelNo, suruculinsansNo)

            # Sonuçları al
            rows = cursor.fetchall()

            # TableWidget'i temizle
            self.surucuSorgulama_tableWidget.setRowCount(0)

            # Sonuçları TableWidget'e ekle
            for row_number, row_data in enumerate(rows):
                self.surucuSorgulama_tableWidget.insertRow(row_number)
                for column_number, data in enumerate(row_data):
                    item = QTableWidgetItem(str(data) if data is not None else "")
                    self.surucuSorgulama_tableWidget.setItem(row_number, column_number, item)

        except pyodbc.Error as ex:
            print("Veritabanı Hatası:", ex)
            QMessageBox.critical(self, "Hata", "Veritabanı bağlantı hatası!")

        finally:
            conn.close()

    def surucu_filtreleme(self):
        dosya_yolu, _ = QFileDialog.getSaveFileName(
            self, "Excel Dosyasını Kaydet", "", "Excel Dosyası (*.xlsx)"
        )
        
        if not dosya_yolu:
            return  # Dosya seçilmedi

        try:
            # Excel çalışma kitabı oluştur
            wb = Workbook()
            ws = wb.active
            ws.title = "Seferler"

            # TableWidget başlıklarını yaz
            kolon_sayisi = self.surucuSorgulama_tableWidget.columnCount()
            satir_sayisi = self.surucuSorgulama_tableWidget.rowCount()

            # Başlıkları Excel'e ekle
            for kolon in range(kolon_sayisi):
                baslik = self.surucuSorgulama_tableWidget.horizontalHeaderItem(kolon).text()
                ws.cell(row=1, column=kolon + 1, value=baslik)

            # TableWidget verilerini Excel'e ekle
            for satir in range(satir_sayisi):
                for kolon in range(kolon_sayisi):
                    hucre_verisi = self.surucuSorgulama_tableWidget.item(satir, kolon)
                    ws.cell(row=satir + 2, column=kolon + 1, value=hucre_verisi.text() if hucre_verisi else "")

            # Dosyayı kaydet
            wb.save(dosya_yolu)
            QMessageBox.information(self, "Başarılı", "Veriler Excel dosyasına başarıyla aktarıldı!")
        
        except Exception as ex:
            QMessageBox.critical(self, "Hata", f"Dosya kaydedilirken hata oluştu!\n{ex}")

    def silinecek_arac_musteri_surucu_getir_sql(self):
        try:
            conn = pyodbc.connect(conn_string)
            cursor = conn.cursor()
            '''
            self.comboboxSilinecekSefer.addItem(" ")
            self.comboboxSilinecekArac.addItem(" ")
            self.comboboxSilinecekMusteri.addItem(" ")
            self.comboboxSilinecekSurucu.addItem(" ")
            '''
            
            cursor.execute("SELECT journeyID, departureCity, arrivalCity, departureDate, arrivalDate, busID, driverID FROM Journey")
            for row in cursor.fetchall():
                self.comboboxSilinecekSefer.addItem(f"{row.journeyID} Kalkış: {row.departureCity} | Varış: {row.arrivalCity} | Kalkış Saati: {row.departureDate} | Varış Saati: {row.arrivalDate}") 
            
            cursor.execute("SELECT busID, capacity, numberPlate, marka, bustypeID, model FROM Bus")
            for row in cursor.fetchall():
                self.comboboxSilinecekArac.addItem(f"{row.busID} | Kapasite: {row.capacity} | Plaka: {row.numberPlate} | Marka: {row.marka} | Modeli: {row.model}") 

            cursor.execute("SELECT driverID, firstName, lastName, phone, licenseNumber FROM Driver")
            for row in cursor.fetchall():
                self.comboboxSilinecekSurucu.addItem(f"{row.driverID} | İsim Soyisim: {row.firstName} {row.lastName} | Telefon Numarası: {row.phone} | Ehliyet Numarası: {row.licenseNumber}") 
            

            cursor.execute("SELECT customerID, firstName, lastName, phone, email, tcKimlikNo, authorizationID FROM Customer")
            for row in cursor.fetchall():
                self.comboboxSilinecekMusteri.addItem(f"{row.customerID} | İsim Soyisim: {row.firstName} {row.lastName} | Telefon Numarası: {row.phone} | E-mail: {row.email} | TC Kimlik No: {row.tcKimlikNo}")

        except pyodbc.Error as ex:
            print(ex)

        finally:
            conn.close()
        
    def sefer_silme(self):
        SilinecekSefer = self.comboboxSilinecekSefer.currentText()
        journeyID = SilinecekSefer.split(" ", 1)[0]

        if self.onayCheckBoxSefer.isChecked() and self.comboboxSilinecekSefer.currentText() != " ":
            response = QMessageBox.question(self, "Koşul Sağlandı", "Seferi silmek istediğinizden emin misiniz?",
                                            QMessageBox.Yes | QMessageBox.No)
            if response == QMessageBox.Yes:
                try:
                    conn = pyodbc.connect(conn_string)
                    cursor = conn.cursor()
                    cursor.execute("DELETE FROM Journey WHERE journeyID = ?",(journeyID,))
                    conn.commit()
                    QMessageBox.information(self, "Başarılı", "Araç Silinmiştir")
                    
                except pyodbc.Error as ex:
                    QMessageBox.warning(self, "Hata", "Veritabanında hata oldu tekrar deneyiniz")

                finally:
                    conn.close()
                
        else:
            QMessageBox.warning(self, "Hata", "Seçim yapınız ve Kutucuğu işaretleyiniz")

    def musteri_silme(self):
        SilinecekMusteri = self.comboboxSilinecekMusteri.currentText()
        customerID = SilinecekMusteri.split(" ", 1)[0]

        if self.onayCheckBoxMusteri.isChecked() and self.comboboxSilinecekMusteri.currentText() != " ":
            response = QMessageBox.question(self, "Koşul Sağlandı", "Müşteriyi silmek istediğinizden emin misiniz?",
                                            QMessageBox.Yes | QMessageBox.No)
            if response == QMessageBox.Yes:
                try:
                    conn = pyodbc.connect(conn_string)
                    cursor = conn.cursor()
                    cursor.execute("DELETE FROM Customer WHERE customerID = ?",(customerID,))
                    conn.commit()
                    QMessageBox.information(self, "Başarılı", "Müşteri Silinmiştir")
                    
                except pyodbc.Error as ex:
                    QMessageBox.warning(self, "Hata", "Veritabanında hata oldu tekrar deneyiniz")

                finally:
                    conn.close()
                
        else:
            QMessageBox.warning(self, "Hata", "Seçim yapınız ve Kutucuğu işaretleyiniz")

    def surucu_silme(self):
        SilinecekSurucu = self.comboboxSilinecekSurucu.currentText()
        driverID = SilinecekSurucu.split(" ", 1)[0]

        if self.onayCheckBoxSurucu.isChecked() and self.comboboxSilinecekSurucu.currentText() != " ":
            response = QMessageBox.question(self, "Koşul Sağlandı", "Sürücü silmek istediğinizden emin misiniz?",
                                            QMessageBox.Yes | QMessageBox.No)
            if response == QMessageBox.Yes:
                try:
                    conn = pyodbc.connect(conn_string)
                    cursor = conn.cursor()
                    cursor.execute("DELETE FROM Driver WHERE driverID = ?",(driverID,))
                    conn.commit()
                    QMessageBox.information(self, "Başarılı", "Sürücü Silinmiştir")
                    
                except pyodbc.Error as ex:
                    QMessageBox.warning(self, "Hata", "Veritabanında hata oldu tekrar deneyiniz")

                finally:
                    conn.close()
                
        else:
            QMessageBox.warning(self, "Hata", "Seçim yapınız ve Kutucuğu işaretleyiniz")

    def driver_excel_veri_aktar(self):
        # Excel dosyasını seç
        dosya_yolu, _ = QFileDialog.getOpenFileName(
            self, "Excel Dosyasını Seç", "", "Excel Dosyası (*.xlsx)"
        )

        if not dosya_yolu:
            return  # Dosya seçilmediyse çıkış yap
        
        try:
            # Excel dosyasını yükle
            wb = load_workbook(dosya_yolu)
            ws = wb.active

            # MSSQL veritabanına bağlan
            conn = pyodbc.connect(conn_string)
            cursor = conn.cursor()

            # Satırları oku ve veritabanına ekle
            for row in ws.iter_rows(min_row=2, values_only=True):  # Başlıkları atla
                firstName, lastName, phone, licenseNumber = row

                # SQL Sorgusu
                query = """
                INSERT INTO Driver (firstName, lastName, phone, licenseNumber)
                VALUES (?, ?, ?, ?)
                """
                params = (firstName, lastName, phone, licenseNumber)
                cursor.execute(query, params)

            conn.commit()  # Veritabanına işlemleri kaydet
            QMessageBox.information(self, "Başarılı", "Veriler başarıyla MSSQL'e aktarıldı!")

        except FileNotFoundError:
            QMessageBox.critical(self, "Hata", "Excel dosyası bulunamadı!")
        except pyodbc.Error as ex:
            QMessageBox.critical(self, "Hata", f"Veritabanı hatası oluştu!\n{ex}")
        except Exception as ex:
            QMessageBox.critical(self, "Hata", f"Bilinmeyen bir hata oluştu!\n{ex}")
        finally:
            conn.close()



if __name__ == "__main__":
    app = QApplication(sys.argv)
    win = Window()
    win.show()

    sys.exit(app.exec())
